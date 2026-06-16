# -*- coding: utf-8 -*-
"""
Async and resumable product import execution.
"""

import json
import logging
import time
import tracemalloc
from concurrent.futures import ThreadPoolExecutor, as_completed
from decimal import Decimal, InvalidOperation
from typing import Any, Callable, Dict, List, Optional, Tuple
from uuid import uuid4

from openpyxl import load_workbook

from app.config import SOURCE_TYPE
from app.models import get_db
from app.services.ali1688_service import ali1688_service
from app.services.import_category_mapping_service import import_category_mapping_service
from app.services.category_import_rule_v2 import category_import_rule_v2
from app.services.import_exclusion_report_service import import_exclusion_report_service
from app.services.product_import_service import PLACEHOLDER, ProductImportService
from app.services.wecom_robot_service import wecom_robot_service

logger = logging.getLogger(__name__)

EXCEL_HEADER_OFFER_ID = '宝贝ID'
EXCEL_HEADER_TITLE = '商品标题'
EXCEL_HEADER_LINK = '宝贝链接'
EXCEL_HEADER_CATEGORY = '类目'
DETAIL_EXCLUSION_HAINAN = 'HAINAN_DELIVERY_FORBIDDEN'
DETAIL_EXCLUSION_MIN_ORDER = 'MIN_ORDER_QUANTITY_UNSUPPORTED'
HAINAN_EXCLUSION_REASON = '海南禁止配送，不满足要求，已排除'
MIN_ORDER_EXCLUSION_REASON_TEMPLATE = '最小起订量为{quantity}，不满足要求，已排除'


class ImportCancelledError(Exception):
    pass


class ProductImportJobService(ProductImportService):
    BATCH_SIZE = 50
    IMPORT_CHUNK_SIZE = 100
    API_BATCH_SIZE = 20
    MAX_WORKERS = 8

    UPSERT_COLUMNS = [
        'offer_id', 'title', 'price', 'cost_price', 'sell_price', 'freight', 'image_url',
        'supplier_name', 'sales_count', 'deliver_days', 'stock', 'sync_status', 'description',
        'attributes', 'images', 'sku_info', 'sku_count', 'supplier_id', 'category_id',
        'category_name', 'erp_category_id', 'erp_category_name', 'offer_url', 'comment_count',
        'month_order_count', 'month_distribution_count', 'tags', 'listed_time', 'shop_name',
        'source_type', 'create_time', 'main_video', 'seven_days_refunds', 'product_type',
        'quality_level', 'reference_price', 'seller_login_id', 'product_sale_info',
        'product_extend_infos', 'sale_limit_address', 'service_capabilities',
        'official_logistics_sku_info', 'product_shipping_info'
    ]

    UPSERT_UPDATE_SQL = '''
        title = VALUES(title),
        price = VALUES(price),
        cost_price = VALUES(cost_price),
        sell_price = VALUES(sell_price),
        freight = VALUES(freight),
        image_url = VALUES(image_url),
        supplier_name = VALUES(supplier_name),
        sales_count = VALUES(sales_count),
        deliver_days = VALUES(deliver_days),
        stock = VALUES(stock),
        sync_status = VALUES(sync_status),
        category_id = IFNULL(NULLIF(VALUES(category_id), ''), category_id),
        category_name = IFNULL(NULLIF(VALUES(category_name), ''), category_name),
        erp_category_id = IFNULL(NULLIF(VALUES(erp_category_id), ''), erp_category_id),
        erp_category_name = IFNULL(NULLIF(VALUES(erp_category_name), ''), erp_category_name),
        offer_url = IFNULL(NULLIF(VALUES(offer_url), ''), offer_url),
        comment_count = VALUES(comment_count),
        month_order_count = VALUES(month_order_count),
        month_distribution_count = VALUES(month_distribution_count),
        tags = IFNULL(NULLIF(VALUES(tags), ''), tags),
        listed_time = IFNULL(NULLIF(VALUES(listed_time), ''), listed_time),
        shop_name = IFNULL(NULLIF(VALUES(shop_name), ''), shop_name),
        source_type = VALUES(source_type),
        create_time = VALUES(create_time),
        description = VALUES(description),
        attributes = VALUES(attributes),
        images = VALUES(images),
        sku_info = VALUES(sku_info),
        sku_count = VALUES(sku_count),
        supplier_id = VALUES(supplier_id),
        main_video = VALUES(main_video),
        seven_days_refunds = VALUES(seven_days_refunds),
        product_type = VALUES(product_type),
        quality_level = VALUES(quality_level),
        reference_price = VALUES(reference_price),
        seller_login_id = VALUES(seller_login_id),
        product_sale_info = VALUES(product_sale_info),
        product_extend_infos = VALUES(product_extend_infos),
        sale_limit_address = VALUES(sale_limit_address),
        service_capabilities = VALUES(service_capabilities),
        official_logistics_sku_info = VALUES(official_logistics_sku_info),
        product_shipping_info = VALUES(product_shipping_info)
    '''

    def run_import_from_excel(
        self,
        file_path: str,
        import_task_id: str = '',
        progress_callback: Optional[Callable[[int, int, str], None]] = None,
        should_cancel: Optional[Callable[[], bool]] = None,
        checkpoint_callback: Optional[Callable[[Dict[str, Any]], None]] = None,
        resume_state: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        start_wall = time.perf_counter()
        start_cpu = time.process_time()
        started_tracing = False
        if not tracemalloc.is_tracing():
            tracemalloc.start()
            started_tracing = True

        metrics = dict((resume_state or {}).get('metrics') or {})
        metrics.setdefault('alerts', [])
        metrics.setdefault('batches', [])
        metrics.setdefault('profile', {})
        metrics['profile'].update({
            'chunk_size': self.IMPORT_CHUNK_SIZE,
            'db_batch_size': self.BATCH_SIZE,
            'max_workers': self.MAX_WORKERS,
            'file_path': file_path,
        })

        rows_data, validation_failures, initial_excluded_rows = self._load_excel_rows_with_category_mapping(file_path, import_task_id=import_task_id)
        total = len(rows_data) + len(validation_failures) + len(initial_excluded_rows)
        if total <= 0:
            raise ValueError('Excel无有效数据行')

        state = dict(resume_state or {})
        checkpoint_offset = max(0, min(int(state.get('checkpoint_offset') or 0), len(rows_data)))
        success_count = int(state.get('success_count') or 0)
        fail_count = int(state.get('fail_count') or 0)
        failed_details = list(state.get('failed_details') or [])
        revision_job_ids = list(state.get('revision_job_ids') or [])
        excluded_rows = list(state.get('excluded_rows') or initial_excluded_rows)
        detail_exclusion_summary = self._normalize_detail_exclusion_summary(state.get('detail_exclusion_summary'))

        if not state:
            failed_details.extend(validation_failures)
            fail_count = len(failed_details)
        excluded_count = len(excluded_rows)
        prechecked_count = len(validation_failures) + len(initial_excluded_rows)

        self._emit_progress(progress_callback, max(1, int(((checkpoint_offset + prechecked_count) / max(total, 1)) * 100)), f'待导入 {total} 条，已校验 {len(validation_failures)} 条异常')

        logger.info(
            'Import pre-check summary total=%s included=%s validation_failures=%s excluded=%s',
            total,
            len(rows_data),
            len(validation_failures),
            excluded_count
        )
        for batch_start in range(checkpoint_offset, len(rows_data), self.IMPORT_CHUNK_SIZE):
            self._raise_if_cancelled(should_cancel)
            batch_rows = rows_data[batch_start:batch_start + self.IMPORT_CHUNK_SIZE]
            batch_offer_ids = [item['offer_id'] for item in batch_rows]
            batch_no = batch_start // self.IMPORT_CHUNK_SIZE + 1
            batch_metric: Dict[str, Any] = {
                'batch_no': batch_no,
                'row_count': len(batch_rows),
                'started_at_ms': int(time.time() * 1000),
            }

            self._emit_progress(
                progress_callback,
                max(1, int(((batch_start + prechecked_count) / max(total, 1)) * 100)),
                f'第 {batch_no} 批：获取商品详情 ({len(batch_rows)} 条)'
            )
            fetch_started = time.perf_counter()
            details_map = self._fetch_product_details(
                batch_offer_ids,
                should_cancel=should_cancel,
                item_progress_callback=lambda completed, subtotal: self._emit_progress(
                    progress_callback,
                    max(1, int((((batch_start + completed) + prechecked_count) / max(total, 1)) * 100)),
                    f'第 {batch_no} 批：获取商品详情 {completed}/{subtotal}'
                )
            )
            batch_metric['fetch_seconds'] = round(time.perf_counter() - fetch_started, 3)

            batch_rows_to_import, batch_detail_excluded_rows = self._filter_rows_by_detail_rules(batch_rows, details_map)
            if batch_detail_excluded_rows:
                excluded_rows.extend(batch_detail_excluded_rows)
                excluded_count = len(excluded_rows)
                self._merge_detail_exclusion_summary(detail_exclusion_summary, batch_detail_excluded_rows)
                details_map = {
                    row['offer_id']: details_map.get(row['offer_id'], {})
                    for row in batch_rows_to_import
                }
                logger.info(
                    'Import detail exclusion batch=%s excluded=%s hainan_total=%s min_order_total=%s',
                    batch_no,
                    len(batch_detail_excluded_rows),
                    detail_exclusion_summary.get('hainan_count', 0),
                    detail_exclusion_summary.get('min_order_quantity_count', 0)
                )

            self._raise_if_cancelled(should_cancel)
            upload_started = time.perf_counter()
            url_mapping = self._upload_images_to_cos(details_map, batch_rows_to_import)
            if url_mapping:
                self._replace_image_urls(details_map, url_mapping, batch_rows_to_import)
            batch_metric['image_seconds'] = round(time.perf_counter() - upload_started, 3)
            batch_metric['image_count'] = len(url_mapping or {})

            self._emit_progress(
                progress_callback,
                max(1, int((((batch_start + len(batch_rows) * 0.8) + prechecked_count) / max(total, 1)) * 100)),
                f'第 {batch_no} 批：准备入库'
            )
            build_started = time.perf_counter()
            products_to_import = []
            batch_failed_details = []
            for row_info in batch_rows_to_import:
                self._raise_if_cancelled(should_cancel)
                row_index = row_info['row_index']
                offer_id = row_info['offer_id']
                detail_result = details_map.get(offer_id, {})
                if not detail_result.get('success'):
                    logger.info(
                        'Import row failed at detail fetch row=%s offer_id=%s reason=%s',
                        row_index,
                        offer_id,
                        detail_result.get('error', 'detail fetch failed')
                    )
                    batch_failed_details.append({
                        'row': row_index,
                        'offer_id': offer_id,
                        'reason': detail_result.get('error', '详情获取失败')
                    })
                    continue
                try:
                    product_data = self._build_product_data(
                        offer_id=offer_id,
                        row_data=row_info['raw_data'],
                        detail_result=detail_result
                    )
                    product_data['row_index'] = row_index
                    product_data['resolved_category'] = row_info.get('resolved_category')
                    products_to_import.append(product_data)
                    logger.info(
                        'Import row prepared row=%s offer_id=%s match_type=%s erp_category=%s score=%s',
                        row_index,
                        offer_id,
                        (row_info.get('resolved_category') or {}).get('match_type', ''),
                        ((row_info.get('resolved_category') or {}).get('category') or {}).get('name', ''),
                        (row_info.get('resolved_category') or {}).get('score', 0)
                    )
                except Exception as exc:
                    logger.info('Import row build failed row=%s offer_id=%s reason=%s', row_index, offer_id, exc)
                    batch_failed_details.append({
                        'row': row_index,
                        'offer_id': offer_id,
                        'reason': str(exc)
                    })
            batch_metric['build_seconds'] = round(time.perf_counter() - build_started, 3)

            self._emit_progress(
                progress_callback,
                max(1, int((((batch_start + len(batch_rows) * 0.9) + prechecked_count) / max(total, 1)) * 100)),
                f'第 {batch_no} 批：写入数据库'
            )
            db_started = time.perf_counter()
            import_result = self._batch_import_to_database(products_to_import)
            batch_metric['db_seconds'] = round(time.perf_counter() - db_started, 3)

            batch_failed_details.extend(import_result.get('failed_details', []))
            failed_details.extend(batch_failed_details)
            success_count += int(import_result.get('success_count') or 0)
            fail_count += len(batch_failed_details)

            revision_job_id = str(import_result.get('revision_job_id') or '').strip()
            if revision_job_id:
                revision_job_ids.append(revision_job_id)

            processed_count = min(total, batch_start + len(batch_rows) + prechecked_count)
            current_memory_mb, peak_memory_mb = self._get_memory_metrics()
            batch_metric.update({
                'processed_count': processed_count,
                'success_count': success_count,
                'fail_count': fail_count,
                'excluded_count': excluded_count,
                'current_memory_mb': current_memory_mb,
                'peak_memory_mb': peak_memory_mb,
                'success_in_batch': int(import_result.get('success_count') or 0),
                'fail_in_batch': len(batch_failed_details),
                'excluded_in_batch': len(batch_detail_excluded_rows),
            })
            self._append_alerts(metrics, batch_metric)
            metrics['batches'].append(batch_metric)
            metrics['batches'] = metrics['batches'][-30:]

            checkpoint_state = {
                'total': total,
                'success_count': success_count,
                'fail_count': fail_count,
                'failed_details': failed_details,
                'checkpoint_offset': batch_start + len(batch_rows),
                'processed_count': processed_count,
                'progress': min(99, int((processed_count / max(total, 1)) * 100)),
                'metrics': metrics,
                'revision_job_ids': revision_job_ids,
                'excluded_rows': excluded_rows,
                'detail_exclusion_summary': detail_exclusion_summary,
            }
            if checkpoint_callback:
                checkpoint_callback(checkpoint_state)

        elapsed_seconds = round(time.perf_counter() - start_wall, 3)
        cpu_seconds = round(time.process_time() - start_cpu, 3)
        current_memory_mb, peak_memory_mb = self._get_memory_metrics()
        metrics['profile'].update({
            'elapsed_seconds': elapsed_seconds,
            'cpu_seconds': cpu_seconds,
            'current_memory_mb': current_memory_mb,
            'peak_memory_mb': peak_memory_mb,
        })

        excluded_report = import_exclusion_report_service.create_report(
            excluded_rows,
            source=SOURCE_TYPE,
            import_task_id=import_task_id
        )
        if excluded_report.get('count'):
            wecom_robot_service.send_import_exclusion_alert({
                'excluded_count': excluded_report.get('count'),
                'report_url': excluded_report.get('report_url'),
                'import_task_id': import_task_id,
            })
        exclusion_summary = self._build_exclusion_summary(
            excluded_rows,
            initial_category_excluded_count=len(initial_excluded_rows),
            detail_exclusion_summary=detail_exclusion_summary,
        )

        has_real_failure = fail_count > 0
        has_exclusion = excluded_count > 0
        result = {
            'success': not has_real_failure,
            'partial_success': has_exclusion or (success_count > 0 and has_real_failure),
            'total': total,
            'success_count': success_count,
            'fail_count': fail_count,
            'excluded_count': excluded_count,
            'excluded_report': excluded_report,
            'exclusion_summary': exclusion_summary,
            'failed_details': failed_details,
            'checkpoint_offset': len(rows_data),
            'processed_count': total,
            'revision_job_ids': revision_job_ids,
            'metrics': metrics,
        }
        if has_real_failure and failed_details:
            top_errors = failed_details[:10]
            error_lines = [f"第{d.get('row','?')}行(offer_id={d.get('offer_id','')}): {d.get('reason','')}" for d in top_errors]
            result['error'] = f"共{fail_count}条导入失败：\n" + "\n".join(error_lines)
            if fail_count > 10:
                result['error'] += f"\n...还有{fail_count - 10}条失败记录"
        if has_exclusion and not has_real_failure:
            summary_msgs = (exclusion_summary or {}).get('messages') or []
            if summary_msgs:
                result['error'] = f"导入完成，{excluded_count}件商品被排除：{'；'.join(summary_msgs)}"
        if not has_real_failure and not has_exclusion:
            result.pop('error', None)
        self._emit_progress(progress_callback, 100, f'导入完成，成功 {success_count} 条，失败 {fail_count} 条')

        if started_tracing:
            tracemalloc.stop()
        return result

    def _load_excel_rows(self, file_path: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
        try:
            sheet = workbook.active
            if sheet.max_row is not None and sheet.max_row < 2:
                raise ValueError('Excel无有效数据行')

            header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(value).strip() if value is not None else '' for value in header_cells]
            required = ['宝贝ID', '商品标题']
            missing = [item for item in required if item not in headers]
            if missing:
                raise ValueError(f"Excel缺少必要列: {','.join(missing)}")

            rows_data: List[Dict[str, Any]] = []
            failed_details: List[Dict[str, Any]] = []
            seen_offer_ids = set()
            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                row_data = {headers[idx]: row[idx] for idx in range(len(headers))}
                offer_id = self._extract_offer_id(row_data.get('宝贝ID') or row_data.get('宝贝链接'))
                title = str(row_data.get('商品标题') or '').strip()
                if not offer_id:
                    failed_details.append({'row': row_index, 'offer_id': '', 'reason': '宝贝ID为空或格式错误'})
                    continue
                if not title:
                    failed_details.append({'row': row_index, 'offer_id': offer_id, 'reason': '商品标题不能为空'})
                    continue
                if offer_id in seen_offer_ids:
                    failed_details.append({'row': row_index, 'offer_id': offer_id, 'reason': 'Excel中存在重复的SPU'})
                    continue
                seen_offer_ids.add(offer_id)
                rows_data.append({
                    'row_index': row_index,
                    'offer_id': offer_id,
                    'raw_data': row_data,
                })
            if not rows_data and failed_details:
                raise ValueError('未找到有效的商品ID')
            return rows_data, failed_details
        finally:
            workbook.close()

    '''
    '''
    '''
    def _load_excel_rows_with_category_mapping(self, file_path: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
        conn = get_db()
        cursor = conn.cursor()
        try:
            category_cache = import_category_mapping_service.load_category_cache(cursor)
        finally:
            conn.close()

        workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
        try:
            sheet = workbook.active
            if sheet.max_row is not None and sheet.max_row < 2:
                raise ValueError('Excel鏃犳湁鏁堟暟鎹')

            header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(value).strip() if value is not None else '' for value in header_cells]
            required = ['瀹濊礉ID', '鍟嗗搧鏍囬']
            missing = [item for item in required if item not in headers]
            if missing:
                raise ValueError(f"Excel缂哄皯蹇呰鍒? {','.join(missing)}")

            rows_data: List[Dict[str, Any]] = []
            failed_details: List[Dict[str, Any]] = []
            excluded_rows: List[Dict[str, Any]] = []
            seen_offer_ids = set()
            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                row_data = {headers[idx]: row[idx] for idx in range(len(headers))}
                offer_id = self._extract_offer_id(row_data.get('瀹濊礉ID') or row_data.get('瀹濊礉閾炬帴'))
                title = str(row_data.get('鍟嗗搧鏍囬') or '').strip()
                raw_category_name = str(row_data.get('绫荤洰') or '').strip()
                if not offer_id:
                    logger.info('Import row invalid row=%s reason=missing_offer_id', row_index)
                    failed_details.append({'row': row_index, 'offer_id': '', 'reason': '瀹濊礉ID涓虹┖鎴栨牸寮忛敊璇?'})
                    continue
                if not title:
                    logger.info('Import row invalid row=%s offer_id=%s reason=missing_title', row_index, offer_id)
                    failed_details.append({'row': row_index, 'offer_id': offer_id, 'reason': '鍟嗗搧鏍囬涓嶈兘涓虹┖'})
                    continue
                if offer_id in seen_offer_ids:
                    logger.info('Import row invalid row=%s offer_id=%s reason=duplicate_offer_id', row_index, offer_id)
                    failed_details.append({'row': row_index, 'offer_id': offer_id, 'reason': 'Excel涓瓨鍦ㄩ噸澶嶇殑SPU'})
                    continue

                mapping_result = import_category_mapping_service.evaluate_row(category_cache, raw_category_name, title)
                if mapping_result.get('excluded'):
                    logger.info(
                        'Import row excluded row=%s offer_id=%s category=%s reason=%s',
                        row_index,
                        offer_id,
                        raw_category_name,
                        mapping_result.get('reason', '')
                    )
                    excluded_rows.append({
                        'row_index': row_index,
                        'raw_category_name': raw_category_name,
                        'exclude_reason': mapping_result.get('reason', ''),
                        'title': title,
                        'suggestion': mapping_result.get('suggestion', ''),
                        'offer_id': offer_id,
                    })
                    continue

                seen_offer_ids.add(offer_id)
                logger.info(
                    'Import row category mapped row=%s offer_id=%s category=%s match_type=%s target=%s score=%s',
                    row_index,
                    offer_id,
                    raw_category_name,
                    mapping_result.get('match_type', ''),
                    (mapping_result.get('category') or {}).get('name', ''),
                    mapping_result.get('score', 0)
                )
                rows_data.append({
                    'row_index': row_index,
                    'offer_id': offer_id,
                    'raw_data': row_data,
                    'resolved_category': mapping_result,
                })
            if not rows_data and failed_details and not excluded_rows:
                raise ValueError('鏈壘鍒版湁鏁堢殑鍟嗗搧ID')
            return rows_data, failed_details, excluded_rows
        finally:
            workbook.close()

    '''

    def _fetch_product_details(
        self,
        offer_ids: List[str],
        should_cancel: Optional[Callable[[], bool]] = None,
        item_progress_callback: Optional[Callable[[int, int], None]] = None
    ) -> Dict[str, Dict[str, Any]]:
        results: Dict[str, Dict[str, Any]] = {}
        total = len(offer_ids)
        if total <= 0:
            return results

        executor = ThreadPoolExecutor(max_workers=self.MAX_WORKERS)
        try:
            future_to_offer_id = {
                executor.submit(ali1688_service.get_distribution_product_info, offer_id): offer_id
                for offer_id in offer_ids
            }
            completed = 0
            for future in as_completed(future_to_offer_id):
                if should_cancel and should_cancel():
                    executor.shutdown(wait=False, cancel_futures=True)
                    raise ImportCancelledError('导入已取消')
                offer_id = future_to_offer_id[future]
                try:
                    result = future.result(timeout=30)
                    if result.get('success'):
                        result['source'] = SOURCE_TYPE
                    results[offer_id] = result
                except Exception as exc:
                    results[offer_id] = {'success': False, 'error': str(exc)}
                completed += 1
                if item_progress_callback:
                    item_progress_callback(completed, total)
        finally:
            executor.shutdown(wait=False, cancel_futures=True)
        return results

    def _filter_rows_by_detail_rules(
        self,
        batch_rows: List[Dict[str, Any]],
        details_map: Dict[str, Dict[str, Any]]
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        rows_to_import: List[Dict[str, Any]] = []
        excluded_rows: List[Dict[str, Any]] = []
        for row_info in batch_rows:
            offer_id = str(row_info.get('offer_id') or '').strip()
            detail_result = details_map.get(offer_id, {})
            if detail_result.get('success'):
                exclusion = self._build_detail_rule_exclusion(row_info, detail_result)
                if exclusion:
                    excluded_rows.append(exclusion)
                    logger.info(
                        'Import row excluded by detail rules row=%s offer_id=%s reason=%s',
                        exclusion.get('row_index'),
                        offer_id,
                        exclusion.get('exclude_reason')
                    )
                    continue
            rows_to_import.append(row_info)
        return rows_to_import, excluded_rows

    def _build_detail_rule_exclusion(
        self,
        row_info: Dict[str, Any],
        detail_result: Dict[str, Any]
    ) -> Optional[Dict[str, Any]]:
        detail_data = detail_result.get('detail') if isinstance(detail_result, dict) else {}
        if not isinstance(detail_data, dict):
            detail_data = {}

        reasons: List[str] = []
        codes: List[str] = []
        suggestion_parts: List[str] = []

        limit_address_codes = self._extract_limit_address_codes(detail_data, detail_result)
        if any(str(code).strip().startswith('46') for code in limit_address_codes):
            reasons.append(HAINAN_EXCLUSION_REASON)
            codes.append(DETAIL_EXCLUSION_HAINAN)
            suggestion_parts.append(f"limitAddressCodes={','.join(limit_address_codes[:20])}")

        min_order_quantity = self._extract_min_order_quantity(detail_data, detail_result)
        if min_order_quantity is not None and min_order_quantity != Decimal('1'):
            display_quantity = self._format_decimal_value(min_order_quantity)
            reasons.append(MIN_ORDER_EXCLUSION_REASON_TEMPLATE.format(quantity=display_quantity))
            codes.append(DETAIL_EXCLUSION_MIN_ORDER)
            suggestion_parts.append(f"minOrderQuantity={display_quantity}")

        if not reasons:
            return None

        row_data = row_info.get('raw_data') or {}
        return {
            'row_index': row_info.get('row_index') or 0,
            'raw_category_name': str(row_data.get(EXCEL_HEADER_CATEGORY) or ''),
            'exclude_reason': '；'.join(reasons),
            'title': str(row_data.get(EXCEL_HEADER_TITLE) or detail_data.get('title') or ''),
            'suggestion': '；'.join(suggestion_parts),
            'offer_id': str(row_info.get('offer_id') or '').strip(),
            'exclusion_codes': codes,
            'limit_address_codes': limit_address_codes,
            'min_order_quantity': self._format_decimal_value(min_order_quantity) if min_order_quantity is not None else '',
        }

    def _extract_limit_address_codes(
        self,
        detail_data: Dict[str, Any],
        detail_result: Dict[str, Any]
    ) -> List[str]:
        sale_limit = self._first_structured_value([
            detail_data.get('sale_limit_address'),
            detail_data.get('saleLimitAddress'),
        ])
        if sale_limit in ({}, [], None):
            product_info = self._extract_product_info_from_raw((detail_result or {}).get('raw'))
            sale_limit = self._first_structured_value([
                product_info.get('saleLimitAddress') if isinstance(product_info, dict) else None,
            ])

        codes_value = None
        if isinstance(sale_limit, dict):
            for key in ('limitAddressCodes', 'limit_address_codes', 'limitAddressCode'):
                if key in sale_limit:
                    codes_value = sale_limit.get(key)
                    break
        elif isinstance(sale_limit, list):
            codes_value = sale_limit

        return self._flatten_scalar_strings(codes_value)

    def _extract_min_order_quantity(
        self,
        detail_data: Dict[str, Any],
        detail_result: Dict[str, Any]
    ) -> Optional[Decimal]:
        sale_info = self._first_structured_value([
            detail_data.get('product_sale_info'),
            detail_data.get('productSaleInfo'),
            detail_data.get('sale_info'),
            detail_data.get('saleInfo'),
        ])
        if sale_info in ({}, [], None):
            product_info = self._extract_product_info_from_raw((detail_result or {}).get('raw'))
            sale_info = self._first_structured_value([
                product_info.get('productSaleInfo') if isinstance(product_info, dict) else None,
                product_info.get('saleInfo') if isinstance(product_info, dict) else None,
            ])

        candidates: List[Any] = []
        if isinstance(sale_info, dict):
            candidates.extend([
                sale_info.get('minOrderQuantity'),
                sale_info.get('min_order_quantity'),
                sale_info.get('minOrder'),
                sale_info.get('min_order'),
            ])
        candidates.extend([
            detail_data.get('min_order'),
            detail_data.get('minOrderQuantity'),
        ])

        for value in candidates:
            amount = self._to_decimal_or_none(value)
            if amount is not None:
                return amount
        return None

    def _extract_product_info_from_raw(self, raw_value: Any) -> Dict[str, Any]:
        raw = self._parse_json_value(raw_value)
        if not isinstance(raw, dict):
            return {}
        if isinstance(raw.get('productInfo'), dict):
            return raw.get('productInfo') or {}
        result = raw.get('result')
        if isinstance(result, dict):
            if isinstance(result.get('productInfo'), dict):
                return result.get('productInfo') or {}
            nested = result.get('result')
            if isinstance(nested, dict):
                if isinstance(nested.get('productInfo'), dict):
                    return nested.get('productInfo') or {}
                return nested
            return result
        return {}

    def _first_structured_value(self, values: List[Any]) -> Any:
        for value in values:
            parsed = self._parse_json_value(value)
            if parsed not in ({}, [], None, ''):
                return parsed
        return {}

    def _parse_json_value(self, value: Any) -> Any:
        if isinstance(value, (dict, list)):
            return value
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return {}
            if text[:1] in ('{', '['):
                try:
                    return json.loads(text)
                except Exception:
                    return value
            return value
        return value

    def _flatten_scalar_strings(self, value: Any) -> List[str]:
        parsed = self._parse_json_value(value)
        if parsed is None:
            return []
        if isinstance(parsed, dict):
            values: List[str] = []
            for key, item in parsed.items():
                values.extend(self._flatten_scalar_strings(key))
                values.extend(self._flatten_scalar_strings(item))
            return values
        if isinstance(parsed, (list, tuple, set)):
            values: List[str] = []
            for item in parsed:
                values.extend(self._flatten_scalar_strings(item))
            return values
        text = str(parsed).strip()
        if not text:
            return []
        if any(separator in text for separator in [',', '，', ';', '；', '\n', '\t', ' ']):
            parts = [
                part.strip()
                for part in text.replace('，', ',').replace('；', ';').replace('\n', ',').replace('\t', ',').replace(' ', ',').replace(';', ',').split(',')
                if part.strip()
            ]
            return parts or [text]
        return [text]

    def _to_decimal_or_none(self, value: Any) -> Optional[Decimal]:
        if value is None or value == '':
            return None
        try:
            return Decimal(str(value).strip())
        except (InvalidOperation, ValueError, TypeError):
            return None

    def _format_decimal_value(self, value: Optional[Decimal]) -> str:
        if value is None:
            return ''
        if value == value.to_integral_value():
            return str(int(value))
        return format(value.normalize(), 'f')

    def _normalize_detail_exclusion_summary(self, summary: Any = None) -> Dict[str, Any]:
        raw = dict(summary or {}) if isinstance(summary, dict) else {}
        value_counts = raw.get('min_order_quantity_values') or raw.get('min_order_quantity_value_counts') or {}
        if isinstance(value_counts, list):
            value_counts = {
                str(item.get('min_order_quantity') or ''): int(item.get('count') or 0)
                for item in value_counts
                if isinstance(item, dict) and str(item.get('min_order_quantity') or '')
            }
        if not isinstance(value_counts, dict):
            value_counts = {}
        return {
            'detail_excluded_count': int(raw.get('detail_excluded_count') or 0),
            'hainan_count': int(raw.get('hainan_count') or 0),
            'min_order_quantity_count': int(raw.get('min_order_quantity_count') or 0),
            'min_order_quantity_value_counts': {
                str(key): int(value or 0)
                for key, value in value_counts.items()
                if str(key)
            },
        }

    def _merge_detail_exclusion_summary(self, summary: Dict[str, Any], excluded_rows: List[Dict[str, Any]]) -> None:
        value_counts = summary.setdefault('min_order_quantity_value_counts', {})
        for item in excluded_rows or []:
            codes = set(item.get('exclusion_codes') or [])
            summary['detail_excluded_count'] = int(summary.get('detail_excluded_count') or 0) + 1
            if DETAIL_EXCLUSION_HAINAN in codes:
                summary['hainan_count'] = int(summary.get('hainan_count') or 0) + 1
            if DETAIL_EXCLUSION_MIN_ORDER in codes:
                summary['min_order_quantity_count'] = int(summary.get('min_order_quantity_count') or 0) + 1
                value = str(item.get('min_order_quantity') or '').strip()
                if value:
                    value_counts[value] = int(value_counts.get(value) or 0) + 1

    def _build_exclusion_summary(
        self,
        excluded_rows: List[Dict[str, Any]],
        initial_category_excluded_count: int,
        detail_exclusion_summary: Dict[str, Any]
    ) -> Dict[str, Any]:
        detail_summary = self._normalize_detail_exclusion_summary(detail_exclusion_summary)
        messages: List[str] = []
        hainan_count = int(detail_summary.get('hainan_count') or 0)
        if hainan_count:
            messages.append(f'{hainan_count}件商品因为海南禁止配送，不满足要求，已排除')

        min_order_values = []
        value_counts = detail_summary.get('min_order_quantity_value_counts') or {}
        for value in sorted(value_counts.keys(), key=lambda item: self._to_decimal_or_none(item) or Decimal('0')):
            count = int(value_counts.get(value) or 0)
            if count <= 0:
                continue
            message = f'{count}件商品因为最小起订量为{value}，不满足要求，已排除'
            messages.append(message)
            min_order_values.append({
                'min_order_quantity': value,
                'count': count,
                'message': message,
            })

        return {
            'total_excluded_count': len(excluded_rows or []),
            'category_excluded_count': int(initial_category_excluded_count or 0),
            'detail_excluded_count': int(detail_summary.get('detail_excluded_count') or 0),
            'hainan_count': hainan_count,
            'min_order_quantity_count': int(detail_summary.get('min_order_quantity_count') or 0),
            'min_order_quantity_values': min_order_values,
            'messages': messages,
        }

    def _batch_import_to_database(self, products: List[Dict[str, Any]]) -> Dict[str, Any]:
        if not products:
            return {'success_count': 0, 'failed_details': [], 'revision_job_id': ''}

        conn = get_db()
        cursor = conn.cursor()
        revision_job_id = uuid4().hex
        success_count = 0
        failed_details: List[Dict[str, Any]] = []
        try:
            existing_product_map = self._load_existing_products(cursor, products)
            leaf_category_cache = self._load_leaf_categories(cursor)
            for start in range(0, len(products), self.BATCH_SIZE):
                batch = products[start:start + self.BATCH_SIZE]
                batch_result = self._execute_batch_insert(
                    conn,
                    cursor,
                    batch,
                    revision_job_id=revision_job_id,
                    existing_product_map=existing_product_map,
                    leaf_category_cache=leaf_category_cache
                )
                success_count += batch_result['success_count']
                failed_details.extend(batch_result['failed_details'])
                if batch_result['price_logs']:
                    self._insert_price_revision_logs(cursor, batch_result['price_logs'], revision_job_id)
                conn.commit()
            return {
                'success_count': success_count,
                'failed_details': failed_details,
                'revision_job_id': revision_job_id,
            }
        except Exception as exc:
            conn.rollback()
            logger.exception('Product import batch failed')
            failed_details.append({'offer_id': '', 'reason': str(exc)})
            return {
                'success_count': success_count,
                'failed_details': failed_details,
                'revision_job_id': revision_job_id,
                'error': str(exc),
            }
        finally:
            conn.close()

    def _execute_batch_insert(
        self,
        conn,
        cursor,
        products: List[Dict[str, Any]],
        revision_job_id: str,
        existing_product_map: Dict[str, Dict[str, Any]],
        leaf_category_cache: Dict[str, Any]
    ) -> Dict[str, Any]:
        prepared_items: List[Dict[str, Any]] = []
        failed_details: List[Dict[str, Any]] = []

        for product in products:
            offer_id = str(product.get('offer_id') or '').strip()
            try:
                existing_product = existing_product_map.get(self._existing_product_key(offer_id, product.get('source')))
                revised_product, price_logs = self._prepare_price_revision(product, existing_product)
                mapped_category = ((revised_product.get('resolved_category') or {}).get('category')) or None
                row_values = self._build_upsert_row(revised_product, mapped_category)
                for item in price_logs:
                    item['revision_job_id'] = revision_job_id
                prepared_items.append({
                    'offer_id': offer_id,
                    'row_index': revised_product.get('row_index'),
                    'row_values': row_values,
                    'price_logs': price_logs,
                    'source': revised_product.get('source', SOURCE_TYPE),
                    'existing': {
                        'offer_id': offer_id,
                        'cost_price': revised_product['cost_price'],
                        'sku_info': revised_product.get('detail_data', {}).get('sku_info', '')
                    }
                })
            except Exception as exc:
                logger.info('Import row database prepare failed row=%s offer_id=%s reason=%s', product.get('row_index'), offer_id, exc)
                failed_details.append({'row': product.get('row_index') or 0, 'offer_id': offer_id, 'reason': str(exc)})

        if not prepared_items:
            return {'success_count': 0, 'failed_details': failed_details, 'price_logs': []}

        price_logs: List[Dict[str, Any]] = []
        try:
            self._bulk_upsert_rows(cursor, [item['row_values'] for item in prepared_items])
            for item in prepared_items:
                existing_product_map[self._existing_product_key(item['offer_id'], item['source'])] = item['existing']
                price_logs.extend(item['price_logs'])
            return {
                'success_count': len(prepared_items),
                'failed_details': failed_details,
                'price_logs': price_logs,
            }
        except Exception as exc:
            logger.warning('Bulk upsert failed, falling back to row mode: %s', exc)
            conn.rollback()

        success_count = 0
        for item in prepared_items:
            try:
                cursor.execute('SAVEPOINT import_row')
                self._bulk_upsert_rows(cursor, [item['row_values']])
                existing_product_map[self._existing_product_key(item['offer_id'], item['source'])] = item['existing']
                price_logs.extend(item['price_logs'])
                success_count += 1
                try:
                    cursor.execute('RELEASE SAVEPOINT import_row')
                except Exception:
                    pass
            except Exception as row_exc:
                try:
                    cursor.execute('ROLLBACK TO SAVEPOINT import_row')
                except Exception:
                    conn.rollback()
                logger.info('Import row database insert failed row=%s offer_id=%s reason=%s', item.get('row_index'), item['offer_id'], row_exc)
                failed_details.append({'row': item.get('row_index') or 0, 'offer_id': item['offer_id'], 'reason': str(row_exc)})
                try:
                    cursor.execute('RELEASE SAVEPOINT import_row')
                except Exception:
                    pass

        return {
            'success_count': success_count,
            'failed_details': failed_details,
            'price_logs': price_logs,
        }

    def _bulk_upsert_rows(self, cursor, rows: List[tuple]) -> None:
        if not rows:
            return
        per_row_sql = '(' + ','.join([PLACEHOLDER] * len(self.UPSERT_COLUMNS)) + ')'
        values_sql = ','.join([per_row_sql] * len(rows))
        flattened_params: List[Any] = []
        for row in rows:
            flattened_params.extend(row)
        cursor.execute(
            f'''
                INSERT INTO import_product
                ({', '.join(self.UPSERT_COLUMNS)})
                VALUES {values_sql}
                ON DUPLICATE KEY UPDATE
                {self.UPSERT_UPDATE_SQL}
            ''',
            flattened_params
        )

    def _build_upsert_row(self, revised_product: Dict[str, Any], mapped_category: Optional[Dict[str, Any]]) -> tuple:
        detail_data = revised_product.get('detail_data') or {}
        erp_category_id = str(mapped_category.get('id')) if mapped_category else ''
        erp_category_name = mapped_category.get('name', '') if mapped_category else ''
        return (
            revised_product['offer_id'],
            revised_product['title'],
            revised_product['price'],
            revised_product['cost_price'],
            revised_product['sell_price'],
            revised_product.get('freight', 0),
            revised_product['image_url'],
            revised_product['supplier_name'],
            revised_product['sales_count'],
            revised_product['deliver_days'],
            revised_product.get('stock', 0),
            'pending',
            detail_data.get('description', ''),
            detail_data.get('attributes', ''),
            detail_data.get('images', ''),
            detail_data.get('sku_info', ''),
            detail_data.get('sku_count', 0),
            detail_data.get('supplier_id', ''),
            detail_data.get('category_id', ''),
            detail_data.get('category_name', ''),
            erp_category_id,
            erp_category_name,
            revised_product['offer_url'],
            revised_product['comment_count'],
            revised_product['month_order_count'],
            revised_product['month_distribution_count'],
            revised_product['tags'],
            revised_product['listed_time'],
            revised_product.get('shop_name', ''),
            str(revised_product.get('source', SOURCE_TYPE)),
            detail_data.get('create_time'),
            detail_data.get('main_video'),
            1 if detail_data.get('seven_days_refunds') else 0,
            detail_data.get('product_type'),
            detail_data.get('quality_level'),
            detail_data.get('reference_price'),
            detail_data.get('seller_login_id'),
            detail_data.get('product_sale_info'),
            detail_data.get('product_extend_infos'),
            detail_data.get('sale_limit_address'),
            detail_data.get('service_capabilities'),
            detail_data.get('official_logistics_sku_info'),
            detail_data.get('product_shipping_info'),
        )

    def _load_existing_products(self, cursor, products: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
        source_values = set()
        offer_ids = set()
        for product in products:
            offer_id = str(product.get('offer_id') or '').strip()
            source = str(product.get('source') or SOURCE_TYPE).strip()
            if offer_id:
                offer_ids.add(offer_id)
                source_values.add(source)
        if not offer_ids:
            return {}
        offer_placeholders = ','.join([PLACEHOLDER] * len(offer_ids))
        source_placeholders = ','.join([PLACEHOLDER] * len(source_values))
        cursor.execute(
            f'''
                SELECT offer_id, source_type, cost_price, sku_info
                FROM import_product
                WHERE offer_id IN ({offer_placeholders})
                  AND source_type IN ({source_placeholders})
            ''',
            list(offer_ids) + list(source_values)
        )
        return {
            self._existing_product_key(row['offer_id'], row.get('source_type')): dict(row)
            for row in cursor.fetchall()
        }

    def _existing_product_key(self, offer_id: Any, source: Any) -> str:
        return f'{str(source or SOURCE_TYPE).strip()}::{str(offer_id or "").strip()}'

    def _load_leaf_categories(self, cursor) -> Dict[str, Any]:
        cursor.execute(
            '''
                SELECT c.id, c.name, c.sort
                FROM erp_category c
                LEFT JOIN erp_category child ON child.parentId = c.id
                WHERE child.id IS NULL
                ORDER BY c.sort ASC, c.id DESC
            '''
        )
        rows = [dict(row) for row in cursor.fetchall()]
        exact_name_map: Dict[str, Dict[str, Any]] = {}
        for row in rows:
            exact_name_map.setdefault(str(row.get('name') or '').strip(), row)
        fuzzy_rows = [row for row in rows if len(str(row.get('name') or '').strip()) >= 2]
        fuzzy_rows.sort(key=lambda item: len(str(item.get('name') or '')), reverse=True)
        return {'exact': exact_name_map, 'fuzzy': fuzzy_rows}

    def _resolve_erp_category_cached(
        self,
        category_cache: Dict[str, Any],
        detail_category_name: str,
        category_name_override: str,
        title: str
    ) -> Optional[Dict[str, Any]]:
        candidates = []
        category_text = str(category_name_override or detail_category_name or '').replace('|', '/').replace('>', '/')
        for part in category_text.split('/'):
            normalized = str(part).strip()
            if normalized:
                candidates.append(normalized)

        exact_map = category_cache.get('exact') or {}
        for name in reversed(candidates):
            row = exact_map.get(name)
            if row:
                return row

        title_text = str(title or '')
        for row in category_cache.get('fuzzy') or []:
            name = str(row.get('name') or '').strip()
            if name and name in title_text:
                return row
        return None

    '''
    def _load_excel_rows_with_category_mapping(self, file_path: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
        conn = get_db()
        cursor = conn.cursor()
        try:
            category_cache = import_category_mapping_service.load_category_cache(cursor)
        finally:
            conn.close()

        workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
        try:
            sheet = workbook.active
            if sheet.max_row is not None and sheet.max_row < 2:
                raise ValueError('Excel无有效数据行')

            header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(value).strip() if value is not None else '' for value in header_cells]
            required = [EXCEL_HEADER_OFFER_ID, EXCEL_HEADER_TITLE]
            missing = [item for item in required if item not in headers]
            if missing:
                raise ValueError(f"Excel缺少必要列: {','.join(missing)}")

            rows_data: List[Dict[str, Any]] = []
            failed_details: List[Dict[str, Any]] = []
            excluded_rows: List[Dict[str, Any]] = []
            seen_offer_ids = set()

            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                row_data = {headers[idx]: row[idx] for idx in range(len(headers))}
                offer_id = self._extract_offer_id(row_data.get(EXCEL_HEADER_OFFER_ID) or row_data.get(EXCEL_HEADER_LINK))
                title = str(row_data.get(EXCEL_HEADER_TITLE) or '').strip()
                raw_category_name = str(row_data.get(EXCEL_HEADER_CATEGORY) or '').strip()

                if not offer_id:
                    logger.info('Import row invalid row=%s reason=missing_offer_id', row_index)
                    failed_details.append({'row': row_index, 'offer_id': '', 'reason': '宝贝ID为空或格式错误'})
                    continue
                if not title:
                    logger.info('Import row invalid row=%s offer_id=%s reason=missing_title', row_index, offer_id)
                    failed_details.append({'row': row_index, 'offer_id': offer_id, 'reason': '商品标题不能为空'})
                    continue
                if offer_id in seen_offer_ids:
                    logger.info('Import row invalid row=%s offer_id=%s reason=duplicate_offer_id', row_index, offer_id)
                    failed_details.append({'row': row_index, 'offer_id': offer_id, 'reason': 'Excel中存在重复的SPU'})
                    continue

                mapping_result = import_category_mapping_service.evaluate_row(category_cache, raw_category_name, title)
                if mapping_result.get('excluded'):
                    logger.info(
                        'Import row excluded row=%s offer_id=%s category=%s reason=%s',
                        row_index,
                        offer_id,
                        raw_category_name,
                        mapping_result.get('reason', '')
                    )
                    excluded_rows.append({
                        'row_index': row_index,
                        'raw_category_name': raw_category_name,
                        'exclude_reason': mapping_result.get('reason', ''),
                        'title': title,
                        'suggestion': mapping_result.get('suggestion', ''),
                        'offer_id': offer_id,
                    })
                    continue

                seen_offer_ids.add(offer_id)
                logger.info(
                    'Import row category mapped row=%s offer_id=%s category=%s match_type=%s target=%s score=%s',
                    row_index,
                    offer_id,
                    raw_category_name,
                    mapping_result.get('match_type', ''),
                    (mapping_result.get('category') or {}).get('name', ''),
                    mapping_result.get('score', 0)
                )
                rows_data.append({
                    'row_index': row_index,
                    'offer_id': offer_id,
                    'raw_data': row_data,
                    'resolved_category': mapping_result,
                })

            if not rows_data and failed_details and not excluded_rows:
                raise ValueError('未找到有效的商品ID')
            return rows_data, failed_details, excluded_rows
        finally:
            workbook.close()

    def _raise_if_cancelled(self, should_cancel: Optional[Callable[[], bool]]) -> None:
        if should_cancel and should_cancel():
            raise ImportCancelledError('导入已取消')

    def _raise_if_cancelled(self, should_cancel: Optional[Callable[[], bool]]) -> None:
        if should_cancel and should_cancel():
            raise ImportCancelledError('导入已取消')

    '''

    '''
    def _load_excel_rows_with_category_mapping(self, file_path: str) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
        conn = get_db()
        cursor = conn.cursor()
        try:
            category_cache = import_category_mapping_service.load_category_cache(cursor)
        finally:
            conn.close()

        workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
        try:
            sheet = workbook.active
            if sheet.max_row is not None and sheet.max_row < 2:
                raise ValueError('Excel无有效数据行')

            header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(value).strip() if value is not None else '' for value in header_cells]
            required = [EXCEL_HEADER_OFFER_ID, EXCEL_HEADER_TITLE]
            missing = [item for item in required if item not in headers]
            if missing:
                raise ValueError(f"Excel缺少必要列: {','.join(missing)}")

            rows_data: List[Dict[str, Any]] = []
            failed_details: List[Dict[str, Any]] = []
            excluded_rows: List[Dict[str, Any]] = []
            seen_offer_ids = set()

            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                row_data = {headers[idx]: row[idx] for idx in range(len(headers))}
                offer_id = self._extract_offer_id(row_data.get(EXCEL_HEADER_OFFER_ID) or row_data.get(EXCEL_HEADER_LINK))
                title = str(row_data.get(EXCEL_HEADER_TITLE) or '').strip()
                raw_category_name = str(row_data.get(EXCEL_HEADER_CATEGORY) or '').strip()

                if not offer_id:
                    logger.info('Import row invalid row=%s reason=missing_offer_id', row_index)
                    failed_details.append({'row': row_index, 'offer_id': '', 'reason': '宝贝ID为空或格式错误'})
                    continue
                if not title:
                    logger.info('Import row invalid row=%s offer_id=%s reason=missing_title', row_index, offer_id)
                    failed_details.append({'row': row_index, 'offer_id': offer_id, 'reason': '商品标题不能为空'})
                    continue
                if offer_id in seen_offer_ids:
                    logger.info('Import row invalid row=%s offer_id=%s reason=duplicate_offer_id', row_index, offer_id)
                    failed_details.append({'row': row_index, 'offer_id': offer_id, 'reason': 'Excel中存在重复的SPU'})
                    continue

                mapping_result = import_category_mapping_service.evaluate_row(category_cache, raw_category_name, title)
                if mapping_result.get('excluded'):
                    logger.info(
                        'Import row excluded row=%s offer_id=%s category=%s reason=%s',
                        row_index,
                        offer_id,
                        raw_category_name,
                        mapping_result.get('reason', '')
                    )
                    excluded_rows.append({
                        'row_index': row_index,
                        'raw_category_name': raw_category_name,
                        'exclude_reason': mapping_result.get('reason', ''),
                        'title': title,
                        'suggestion': mapping_result.get('suggestion', ''),
                        'offer_id': offer_id,
                    })
                    continue

                seen_offer_ids.add(offer_id)
                logger.info(
                    'Import row category mapped row=%s offer_id=%s category=%s match_type=%s target=%s score=%s',
                    row_index,
                    offer_id,
                    raw_category_name,
                    mapping_result.get('match_type', ''),
                    (mapping_result.get('category') or {}).get('name', ''),
                    mapping_result.get('score', 0)
                )
                rows_data.append({
                    'row_index': row_index,
                    'offer_id': offer_id,
                    'raw_data': row_data,
                    'resolved_category': mapping_result,
                })

            if not rows_data and failed_details and not excluded_rows:
                raise ValueError('未找到有效的商品ID')
            return rows_data, failed_details, excluded_rows
        finally:
            workbook.close()

    def _raise_if_cancelled(self, should_cancel: Optional[Callable[[], bool]]) -> None:
        if should_cancel and should_cancel():
            raise ImportCancelledError('导入已取消')

    '''

    def _load_excel_rows_with_category_mapping(
        self,
        file_path: str,
        import_task_id: str = '',
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        读取 Excel，对每行执行 CategoryImportRuleV2 三级类目校验：
        1. 一级类目前置拦截
        2. 全路径精准匹配（O(1)）
        3. 标题相似度兜底（阈值 0.85，500ms 超时）
        被拒绝的行会在方法末尾批量写入 import_log 表。
        """
        conn = get_db()
        cursor = conn.cursor()
        try:
            category_cache = category_import_rule_v2.load_category_cache_v2(cursor)
        finally:
            conn.close()

        workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
        log_entries: List[Dict[str, Any]] = []
        try:
            sheet = workbook.active
            if sheet.max_row is not None and sheet.max_row < 2:
                raise ValueError('Excel无有效数据行')

            header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(value).strip() if value is not None else '' for value in header_cells]
            required = [EXCEL_HEADER_OFFER_ID, EXCEL_HEADER_TITLE]
            missing = [item for item in required if item not in headers]
            if missing:
                raise ValueError(f"Excel缺少必要列: {','.join(missing)}")

            rows_data: List[Dict[str, Any]] = []
            failed_details: List[Dict[str, Any]] = []
            excluded_rows: List[Dict[str, Any]] = []
            seen_offer_ids = set()

            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                row_data = {headers[idx]: row[idx] for idx in range(len(headers))}
                offer_id = self._extract_offer_id(row_data.get(EXCEL_HEADER_OFFER_ID) or row_data.get(EXCEL_HEADER_LINK))
                title = str(row_data.get(EXCEL_HEADER_TITLE) or '').strip()
                raw_category_name = str(row_data.get(EXCEL_HEADER_CATEGORY) or '').strip()

                if not offer_id:
                    logger.info('Import row invalid row=%s reason=missing_offer_id', row_index)
                    failed_details.append({'row': row_index, 'offer_id': '', 'reason': '宝贝ID为空或格式错误'})
                    continue
                if not title:
                    logger.info('Import row invalid row=%s offer_id=%s reason=missing_title', row_index, offer_id)
                    failed_details.append({'row': row_index, 'offer_id': offer_id, 'reason': '商品标题不能为空'})
                    continue
                if offer_id in seen_offer_ids:
                    logger.info('Import row invalid row=%s offer_id=%s reason=duplicate_offer_id', row_index, offer_id)
                    failed_details.append({'row': row_index, 'offer_id': offer_id, 'reason': 'Excel中存在重复的SPU'})
                    continue

                mapping_result = category_import_rule_v2.evaluate_row_v2(
                    category_cache, raw_category_name, title,
                    offer_id=offer_id, import_task_id=import_task_id,
                )
                if mapping_result.get('excluded'):
                    logger.info(
                        'Import row excluded row=%s offer_id=%s category=%s reason=%s',
                        row_index, offer_id, raw_category_name,
                        mapping_result.get('reason', ''),
                    )
                    excluded_rows.append({
                        'row_index': row_index,
                        'raw_category_name': raw_category_name,
                        'exclude_reason': mapping_result.get('reason', ''),
                        'title': title,
                        'suggestion': mapping_result.get('suggestion', ''),
                        'offer_id': offer_id,
                    })
                    log_entry = mapping_result.get('log_entry')
                    if log_entry:
                        log_entries.append(log_entry)
                    continue

                seen_offer_ids.add(offer_id)
                logger.info(
                    'Import row category mapped row=%s offer_id=%s category=%s match_type=%s target=%s score=%s',
                    row_index, offer_id, raw_category_name,
                    mapping_result.get('match_type', ''),
                    (mapping_result.get('category') or {}).get('name', ''),
                    mapping_result.get('score', 0),
                )
                rows_data.append({
                    'row_index': row_index,
                    'offer_id': offer_id,
                    'raw_data': row_data,
                    'resolved_category': mapping_result,
                })

            if not rows_data and failed_details and not excluded_rows:
                raise ValueError('未找到有效的商品ID')
            return rows_data, failed_details, excluded_rows
        finally:
            workbook.close()
            if log_entries:
                try:
                    log_conn = get_db()
                    log_cursor = log_conn.cursor()
                    try:
                        category_import_rule_v2.write_import_logs(log_cursor, log_entries)
                        log_conn.commit()
                    finally:
                        log_conn.close()
                except Exception as exc:
                    logger.warning('Failed to write import_log entries: %s', exc)

    def _raise_if_cancelled(self, should_cancel: Optional[Callable[[], bool]]) -> None:
        if should_cancel and should_cancel():
            raise ImportCancelledError('导入已取消')

    def _emit_progress(self, callback: Optional[Callable[[int, int, str], None]], progress: int, message: str) -> None:
        if callback:
            callback(max(0, min(100, int(progress))), 100, message or '')

    def _get_memory_metrics(self) -> Tuple[float, float]:
        current_bytes, peak_bytes = tracemalloc.get_traced_memory() if tracemalloc.is_tracing() else (0, 0)
        return round(current_bytes / 1024 / 1024, 2), round(peak_bytes / 1024 / 1024, 2)

    def _append_alerts(self, metrics: Dict[str, Any], batch_metric: Dict[str, Any]) -> None:
        alerts = metrics.setdefault('alerts', [])
        if batch_metric.get('db_seconds', 0) >= 5:
            alerts.append(f"batch {batch_metric.get('batch_no')} db stage slow: {batch_metric.get('db_seconds')}s")
        if batch_metric.get('peak_memory_mb', 0) >= 512:
            alerts.append(f"batch {batch_metric.get('batch_no')} memory peak high: {batch_metric.get('peak_memory_mb')}MB")
        if len(alerts) > 20:
            del alerts[:-20]


product_import_job_service = ProductImportJobService()
