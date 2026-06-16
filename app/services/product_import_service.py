# -*- coding: utf-8 -*-
"""
商品导入服务 - 优化版
支持批量API调用、并发处理、批量数据库操作
"""

from app.models import get_db
from app.services.ali1688_service import ali1688_service
from app.config import SOURCE_TYPE
from datetime import datetime
import json
import logging
import re
import threading
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import List, Dict, Any, Optional, Callable
import queue
from uuid import uuid4

PLACEHOLDER = '%s'
logger = logging.getLogger(__name__)


class ProductImportService:
    """优化的商品导入服务"""
    
    # 批量大小配置
    BATCH_SIZE = 50  # 每批处理的商品数量
    API_BATCH_SIZE = 20  # API批量查询的最大商品数
    MAX_WORKERS = 10  # 最大并发线程数
    DESCRIPTION_IMAGE_TEMPLATE = '<p><img src="{url}" alt="" data-href="" width="" height="" style="width: 100%;"/></p>'
    
    def __init__(self):
        self._lock = threading.Lock()
        self._progress_callback: Optional[Callable] = None
    
    def set_progress_callback(self, callback: Callable[[int, int, str], None]):
        """设置进度回调函数
        
        Args:
            callback: 回调函数，参数为 (当前进度, 总数, 状态信息)
        """
        self._progress_callback = callback
    
    def _report_progress(self, current: int, total: int, message: str = ''):
        """报告进度"""
        if self._progress_callback:
            try:
                self._progress_callback(current, total, message)
            except Exception:
                pass
    
    def import_products_from_excel_optimized(
        self, 
        file_stream,
        progress_callback: Optional[Callable[[int, int, str], None]] = None
    ) -> Dict[str, Any]:
        """优化的Excel导入方法
        
        优化点：
        1. 批量获取商品详情（减少API调用次数）
        2. 并发处理（多线程并行）
        3. 批量数据库操作（减少连接开销）
        
        Args:
            file_stream: Excel文件流
            progress_callback: 进度回调函数 (current, total, message)
            
        Returns:
            导入结果统计
        """
        from openpyxl import load_workbook
        import io
        
        self._progress_callback = progress_callback
        self._erp_category_cache.clear()
        self._erp_category_title_cache.clear()
        
        # 1. 读取Excel数据
        self._report_progress(0, 100, '正在读取Excel文件...')
        
        # 将文件流读取到内存中，避免 SpooledTemporaryFile 的问题
        if hasattr(file_stream, 'read'):
            file_content = file_stream.read()
            if isinstance(file_content, bytes):
                file_stream = io.BytesIO(file_content)
            else:
                file_stream = io.StringIO(file_content)
        
        workbook = load_workbook(filename=file_stream, data_only=True)
        sheet = workbook.active
        
        if sheet.max_row is not None and sheet.max_row < 2:
            return {'success': False, 'error': 'Excel无有效数据行'}
        
        # 2. 解析表头
        headers = [str(cell.value).strip() if cell.value is not None else '' for cell in sheet[1]]
        required = ['宝贝ID', '商品标题']
        missing = [item for item in required if item not in headers]
        if missing:
            return {'success': False, 'error': f"Excel缺少必要列: {','.join(missing)}"}
        
        # 3. 解析所有行数据
        rows_data = []
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_data = {headers[idx]: row[idx] for idx in range(len(headers))}
            offer_id = self._extract_offer_id(row_data.get('宝贝ID') or row_data.get('宝贝链接'))
            if offer_id:
                rows_data.append({
                    'row_index': row_index,
                    'offer_id': offer_id,
                    'raw_data': row_data
                })
        
        total = len(rows_data)
        if total == 0:
            return {'success': False, 'error': '未找到有效的商品ID'}
        
        self._report_progress(5, 100, f'共 {total} 件商品待导入...')
        
        # 4. 批量获取商品详情（优化核心）
        offer_ids = [r['offer_id'] for r in rows_data]
        print(f"[Import] Total offer_ids to fetch: {len(offer_ids)}")
        details_map = self._batch_fetch_product_details(offer_ids)
        print(f"[Import] details_map returned: {len(details_map)} items")
        
        # 4.5 上传图片到腾讯云COS
        self._report_progress(55, 100, '正在上传图片到COS...')
        url_mapping = self._upload_images_to_cos(details_map, rows_data)
        if url_mapping:
            self._replace_image_urls(details_map, url_mapping, rows_data)
            print(f"[Import] Images replaced: {len(url_mapping)} URLs")
        
        # 5. 准备导入数据
        self._report_progress(65, 100, '正在准备导入数据...')
        products_to_import = []
        failed_details = []
        
        for row_info in rows_data:
            row_index = row_info['row_index']
            offer_id = row_info['offer_id']
            row_data = row_info['raw_data']
            
            detail_result = details_map.get(offer_id, {})
            
            print(f"[Import] Processing row {row_index}, offer_id={offer_id}, success={detail_result.get('success')}, source={detail_result.get('source')}")
            
            if not detail_result.get('success'):
                failed_details.append({
                    'row': row_index, 
                    'offer_id': offer_id, 
                    'reason': detail_result.get('error', '详情获取失败')
                })
                print(f"[Import] Row {row_index} failed: {detail_result.get('error', '详情获取失败')}")
                continue
            
            # 构建商品数据
            product_data = self._build_product_data(
                offer_id=offer_id,
                row_data=row_data,
                detail_result=detail_result
            )
            products_to_import.append(product_data)
            print(f"[Import] Row {row_index} added to import list, source={product_data.get('source')}")
        
        print(f"[Import] products_to_import count: {len(products_to_import)}, failed_details count: {len(failed_details)}")
        
        # 6. 批量导入数据库
        self._report_progress(75, 100, f'正在导入数据库 ({len(products_to_import)} 件商品)...')
        import_result = self._batch_import_to_database(products_to_import)
        
        # 7. 合并失败记录
        failed_details.extend(import_result.get('failed_details', []))
        
        # 8. 统计结果
        success_count = import_result['success_count']
        fail_count = total - success_count
        
        self._report_progress(100, 100, f'导入完成！成功 {success_count} 件，失败 {fail_count} 件')
        
        return {
            'success': fail_count == 0,
            'partial_success': fail_count > 0 and success_count > 0,
            'total': total,
            'success_count': success_count,
            'fail_count': fail_count,
            'failed_details': failed_details
        }
    
    def _batch_fetch_product_details(self, offer_ids: List[str]) -> Dict[str, Dict]:
        """批量获取商品详情

        策略：全部商品并发查询代发接口，source统一为 'ALIBABA_1688'

        Args:
            offer_ids: 商品ID列表

        Returns:
            {offer_id: detail_result}
        """
        results = {}
        total = len(offer_ids)
        completed = 0

        self._report_progress(10, 100, f'正在获取商品详情 ({total} 件)...')

        with ThreadPoolExecutor(max_workers=self.MAX_WORKERS) as executor:
            # 提交所有任务，直接调用代发接口
            future_to_id = {
                executor.submit(ali1688_service.get_distribution_product_info, offer_id): offer_id
                for offer_id in offer_ids
            }

            # 收集结果
            for future in as_completed(future_to_id):
                offer_id = future_to_id[future]
                try:
                    result = future.result(timeout=30)
                    # 统一标记来源
                    if result.get('success'):
                        result['source'] = SOURCE_TYPE
                    results[offer_id] = result
                except Exception as e:
                    results[offer_id] = {'success': False, 'error': str(e)}

                completed += 1
                progress = 10 + int((completed / total) * 50)
                self._report_progress(progress, 100, f'正在获取商品详情 ({completed}/{total})...')

        return results
    
    def _normalize_image_url(self, url: str) -> str:
        """标准化图片URL，补全域名
        
        Args:
            url: 原始图片URL
            
        Returns:
            完整的图片URL
        """
        if not url:
            return ''
        
        url = str(url).strip()
        if not url:
            return ''
        
        # 已经是完整URL
        if url.startswith('http://') or url.startswith('https://'):
            return url
        
        # 协议相对URL
        if url.startswith('//'):
            return 'https:' + url
        
        # 阿里云CDN相对路径，补全域名
        if url.startswith('img/') or url.startswith('/img/'):
            return 'https://cbu01.alicdn.com/' + url.lstrip('/')
        
        # 其他相对路径
        if not url.startswith('/'):
            return 'https://cbu01.alicdn.com/' + url
        
        return 'https://cbu01.alicdn.com' + url

    def _extract_img_src_from_html(self, html: str) -> List[str]:
        if not html or not isinstance(html, str):
            return []
        pattern = re.compile(r'<img[^>]*\ssrc\s*=\s*["\']?([^"\'>\s]+)', re.IGNORECASE)
        urls = []
        seen = set()
        for match in pattern.findall(html):
            src = str(match).strip()
            if not src or src in seen:
                continue
            seen.add(src)
            urls.append(src)
        return urls

    def _build_description_html_with_template(self, image_urls: List[str]) -> str:
        if not image_urls:
            return ''
        return ''.join([self.DESCRIPTION_IMAGE_TEMPLATE.format(url=url) for url in image_urls if isinstance(url, str) and url.strip()])

    def _upload_images_to_cos(
        self, 
        details_map: Dict[str, Dict],
        rows_data: List[Dict],
        progress_callback: Optional[callable] = None
    ) -> Dict[str, str]:
        """上传图片到腾讯云COS
        
        Args:
            details_map: 商品详情映射 {offer_id: detail_result}
            rows_data: Excel行数据列表
            progress_callback: 进度回调
            
        Returns:
            {原URL: 新URL} 映射
        """
        from app.services.tencent_cos_service import tencent_cos_uploader
        
        if not tencent_cos_uploader.config.is_configured():
            print("[Import] COS未配置，跳过图片上传")
            return {}
        
        # 收集所有图片URL（原始URL -> 标准化URL 映射）
        url_mapping_original = {}  # {原始URL: 标准化URL}
        
        # 从Excel行数据中收集图片地址
        for row_info in rows_data:
            row_data = row_info.get('raw_data', {})
            excel_image_url = str(row_data.get('图片地址') or '').strip()
            if excel_image_url:
                normalized = self._normalize_image_url(excel_image_url)
                if normalized:
                    url_mapping_original[excel_image_url] = normalized
        
        # 从商品详情中收集图片
        for offer_id, detail_result in details_map.items():
            if not detail_result.get('success'):
                continue
            detail_data = detail_result.get('detail', {})
            if not isinstance(detail_data, dict):
                continue
            
            # 主图
            image_url = detail_data.get('image_url', '')
            if image_url:
                normalized = self._normalize_image_url(image_url)
                if normalized:
                    url_mapping_original[image_url] = normalized
            
            # 图片列表
            images = detail_data.get('images', '')
            if images:
                if isinstance(images, str):
                    try:
                        images_list = json.loads(images)
                    except:
                        images_list = []
                elif isinstance(images, list):
                    images_list = images
                else:
                    images_list = []
                    
                for img in images_list:
                    if isinstance(img, str) and img:
                        normalized = self._normalize_image_url(img)
                        if normalized:
                            url_mapping_original[img] = normalized
            
            # SKU图片 - 从sku_info中提取
            sku_info = detail_data.get('sku_info', '')
            if sku_info:
                if isinstance(sku_info, str):
                    try:
                        sku_list = json.loads(sku_info)
                    except:
                        sku_list = []
                elif isinstance(sku_info, list):
                    sku_list = sku_info
                else:
                    sku_list = []
                
                for sku in sku_list:
                    if not isinstance(sku, dict):
                        continue
                    # SKU级别的图片
                    for img_key in ['sku_image_url', 'skuImageUrl', 'image_url', 'imageUrl']:
                        sku_img = sku.get(img_key)
                        if sku_img and isinstance(sku_img, str):
                            normalized = self._normalize_image_url(sku_img)
                            if normalized:
                                url_mapping_original[sku_img] = normalized
                    
                    # Attribute级别的图片
                    attrs = sku.get('attributes', [])
                    if isinstance(attrs, list):
                        for attr in attrs:
                            if not isinstance(attr, dict):
                                continue
                            for attr_img_key in ['sku_image_url', 'skuImageUrl']:
                                attr_img = attr.get(attr_img_key)
                                if attr_img and isinstance(attr_img, str):
                                    normalized = self._normalize_image_url(attr_img)
                                    if normalized:
                                        url_mapping_original[attr_img] = normalized

            # 详情HTML图片 - 从description中提取 img src
            description_html = detail_data.get('description', '')
            if description_html and isinstance(description_html, str):
                desc_img_urls = self._extract_img_src_from_html(description_html)
                for desc_img_url in desc_img_urls:
                    normalized = self._normalize_image_url(desc_img_url)
                    if normalized:
                        url_mapping_original[desc_img_url] = normalized
        
        if not url_mapping_original:
            print("[Import] 没有需要上传的图片")
            return {}
        
        # 获取所有标准化URL（去重）
        normalized_urls = list(set(url_mapping_original.values()))
        print(f"[Import] 开始上传 {len(normalized_urls)} 张图片到COS...")
        
        # 批量上传标准化URL
        normalized_to_cos = tencent_cos_uploader.batch_upload_images(
            normalized_urls,
            prefix='products',
            progress_callback=progress_callback
        )
        
        # 构建原始URL/标准化URL -> COS URL 的映射（双键，提升替换命中率）
        result_mapping = {}
        for original_url, normalized_url in url_mapping_original.items():
            cos_url = normalized_to_cos.get(normalized_url)
            if cos_url and cos_url != normalized_url:
                result_mapping[str(original_url).strip()] = cos_url
                result_mapping[normalized_url] = cos_url
        
        success_count = len(result_mapping)
        failed_normalized = [u for u in normalized_urls if not normalized_to_cos.get(u) or normalized_to_cos.get(u) == u]
        print(f"[Import] 图片上传完成，成功映射 {success_count} 条，上传失败 {len(failed_normalized)} 张")
        if failed_normalized:
            print(f"[Import] 上传失败样例: {failed_normalized[:5]}")
        
        return result_mapping
    
    def _replace_image_urls(
        self,
        details_map: Dict[str, Dict],
        url_mapping: Dict[str, str],
        rows_data: List[Dict]
    ) -> None:
        """替换商品详情中的图片URL
        
        Args:
            details_map: 商品详情映射
            url_mapping: {原URL: 新URL} 映射
            rows_data: Excel行数据列表
        """
        if not url_mapping:
            return

        def map_url(old_url):
            if not old_url or not isinstance(old_url, str):
                return old_url
            stripped = old_url.strip()
            if stripped in url_mapping:
                return url_mapping[stripped]
            normalized = self._normalize_image_url(stripped)
            if normalized and normalized in url_mapping:
                return url_mapping[normalized]
            return old_url
        
        # 替换Excel行数据中的图片地址
        for row_info in rows_data:
            row_data = row_info.get('raw_data', {})
            excel_image_url = str(row_data.get('图片地址') or '').strip()
            if excel_image_url:
                row_data['图片地址'] = map_url(excel_image_url)
        
        # 替换商品详情中的图片URL
        for offer_id, detail_result in details_map.items():
            if not detail_result.get('success'):
                continue
            detail_data = detail_result.get('detail', {})
            if not isinstance(detail_data, dict):
                continue
            
            # 替换主图
            old_image_url = detail_data.get('image_url', '')
            if old_image_url:
                detail_data['image_url'] = map_url(old_image_url)
            
            # 替换图片列表
            images = detail_data.get('images', '')
            if images:
                if isinstance(images, str):
                    try:
                        images_list = json.loads(images)
                    except:
                        images_list = []
                elif isinstance(images, list):
                    images_list = images
                else:
                    images_list = []
                
                new_images_list = []
                for img in images_list:
                    if isinstance(img, str) and img:
                        new_url = map_url(img)
                        new_images_list.append(new_url)
                    else:
                        new_images_list.append(img)
                
                detail_data['images'] = json.dumps(new_images_list, ensure_ascii=False)
            
            # 替换SKU图片
            sku_info = detail_data.get('sku_info', '')
            if sku_info:
                sku_modified = False
                if isinstance(sku_info, str):
                    try:
                        sku_list = json.loads(sku_info)
                    except:
                        sku_list = []
                elif isinstance(sku_info, list):
                    sku_list = sku_info
                else:
                    sku_list = []
                
                for sku in sku_list:
                    if not isinstance(sku, dict):
                        continue
                    
                    # 替换SKU级别的图片
                    for img_key in ['sku_image_url', 'skuImageUrl', 'image_url', 'imageUrl']:
                        old_url = sku.get(img_key)
                        if old_url and isinstance(old_url, str):
                            new_url = map_url(old_url)
                            if new_url != old_url:
                                sku[img_key] = new_url
                                sku_modified = True
                    
                    # 替换Attribute级别的图片
                    attrs = sku.get('attributes', [])
                    if isinstance(attrs, list):
                        for attr in attrs:
                            if not isinstance(attr, dict):
                                continue
                            for attr_img_key in ['sku_image_url', 'skuImageUrl']:
                                old_attr_url = attr.get(attr_img_key)
                                if old_attr_url and isinstance(old_attr_url, str):
                                    new_attr_url = map_url(old_attr_url)
                                    if new_attr_url != old_attr_url:
                                        attr[attr_img_key] = new_attr_url
                                        sku_modified = True
                
                if sku_modified:
                    detail_data['sku_info'] = json.dumps(sku_list, ensure_ascii=False)
                    print(f"[Import] {offer_id}: SKU图片URL已更新")

            # 替换description中的图片并重建HTML模板
            description_html = detail_data.get('description', '')
            if description_html and isinstance(description_html, str):
                original_desc_urls = self._extract_img_src_from_html(description_html)
                rebuilt_urls = []
                for old_desc_url in original_desc_urls:
                    new_desc_url = map_url(old_desc_url)
                    status = 'success' if new_desc_url != old_desc_url else 'fallback'
                    print(f"[Import][Description] offer_id={offer_id} src={old_desc_url} cos={new_desc_url} status={status}")
                    rebuilt_urls.append(new_desc_url)
                rebuilt_html = self._build_description_html_with_template(rebuilt_urls)
                if rebuilt_html:
                    detail_data['description'] = rebuilt_html
                    print(f"[Import] {offer_id}: description图片HTML已重建，图片数={len(rebuilt_urls)}")
            
            # 更新回details_map
            detail_result['detail'] = detail_data
    
    def _build_product_data(
        self, 
        offer_id: str, 
        row_data: Dict, 
        detail_result: Dict
    ) -> Dict:
        """构建商品数据
        
        Args:
            offer_id: 商品ID
            row_data: Excel行数据
            detail_result: API返回的详情结果
            
        Returns:
            完整的商品数据字典
        """
        detail_data = detail_result.get('detail', {})
        if not isinstance(detail_data, dict):
            detail_data = {}
        
        # 从Excel提取基础数据
        category_name = str(row_data.get('类目') or '').strip()
        if category_name and not detail_data.get('category_name'):
            detail_data['category_name'] = category_name
        
        # 直接使用Excel中的原始价格数据，不做自动计算
        price = self._to_float_value(row_data.get('价格'), 0.0)
        sell_price = self._to_float_value(row_data.get('代发价'), 0.0)
        freight = self._to_float_value(
            row_data.get('运费') or row_data.get('邮费') or row_data.get('快递费'),
            0.0
        )
        sku_min_prices = self._apply_sku_min_prices(detail_data, freight)
        sku_price_checked = bool(sku_min_prices)
        if sku_min_prices:
            sell_price = sku_min_prices['min_consign_price']
            price = sku_min_prices['min_market_price']

        return {
            'offer_id': offer_id,
            'title': str(row_data.get('商品标题') or '').strip() or f'商品{offer_id}',
            'price': price,
            'cost_price': price,
            'sell_price': sell_price,
            'freight': freight,
            'image_url': str(row_data.get('图片地址') or '').strip() or detail_data.get('image_url', ''),
            'supplier_name': str(row_data.get('店铺名') or row_data.get('供应商') or '').strip() or detail_data.get('supplier_name', ''),
            'sales_count': self._to_int_value(row_data.get('销量'), 0),
            'deliver_days': self._to_int_value(row_data.get('发货时间'), 48),
            'support_return': '退' in str(row_data.get('标签') or ''),
            'detail_data': detail_data,
            'category_name_override': category_name,
            'offer_url': str(row_data.get('宝贝链接') or '').strip(),
            'comment_count': self._to_int_value(row_data.get('评论数'), 0),
            'month_order_count': self._to_int_value(row_data.get('月成交笔数'), 0),
            'month_distribution_count': self._to_int_value(row_data.get('月代销'), 0),
            'tags': str(row_data.get('标签') or '').strip(),
            'listed_time': str(row_data.get('上架时间') or '').strip(),
            'shop_name': str(row_data.get('店铺') or '').strip(),
            'source': detail_result.get('source', 'yx'),
            'sku_price_checked': sku_price_checked
        }

    def _parse_sku_list_strict(self, sku_info_value, offer_id=''):
        if sku_info_value in (None, ''):
            return []
        if isinstance(sku_info_value, list):
            return sku_info_value
        if isinstance(sku_info_value, str):
            try:
                parsed = json.loads(sku_info_value)
            except Exception as exc:
                raise ValueError(f'offer_id={offer_id} sku_info JSON parse failed: {exc}')
            if isinstance(parsed, list):
                return parsed
        raise ValueError(f'offer_id={offer_id} sku_info must be a JSON array')

    def _price_decimal_optional(self, value) -> Optional[Decimal]:
        if value is None:
            return None
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            value = value.replace(',', '')
        try:
            amount = Decimal(str(value))
        except (InvalidOperation, ValueError, TypeError):
            return None
        if amount <= 0:
            return None
        return amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    def _first_price_decimal(self, sku: Dict, field_names: List[str]) -> Optional[Decimal]:
        for field_name in field_names:
            amount = self._price_decimal_optional(sku.get(field_name))
            if amount is not None:
                return amount
        return None

    def _apply_sku_min_prices(self, detail_data: Dict, freight=0) -> Optional[Dict[str, float]]:
        sku_info_value = detail_data.get('sku_info') if isinstance(detail_data, dict) else None
        if sku_info_value in (None, '', []):
            return None

        sku_list = self._parse_sku_list_strict(sku_info_value)
        if not sku_list:
            return None

        min_consign_price: Optional[Decimal] = None
        min_market_price: Optional[Decimal] = None
        corrected_sku_list = []
        modified = False

        for index, sku in enumerate(sku_list):
            if not isinstance(sku, dict):
                raise ValueError(f'sku_info[{index}] must be an object')

            sku_copy = dict(sku)
            consign_price = self._first_price_decimal(sku_copy, ['consignPrice', 'consign_price', 'price'])
            if consign_price is None:
                sku_id = sku_copy.get('skuId') or sku_copy.get('sku_id') or sku_copy.get('specId') or sku_copy.get('spec_id') or index
                raise ValueError(f'sku price parse failed: missing valid consignPrice for sku={sku_id}')

            market_price = self._first_price_decimal(sku_copy, ['marketPrice', 'market_price'])
            if market_price is None:
                market_price = Decimal(str(self._calculate_sku_market_price(consign_price, 0))).quantize(
                    Decimal('0.01'),
                    rounding=ROUND_HALF_UP
                )
                sku_copy['marketPrice'] = float(market_price)
                modified = True

            min_consign_price = consign_price if min_consign_price is None else min(min_consign_price, consign_price)
            min_market_price = market_price if min_market_price is None else min(min_market_price, market_price)
            corrected_sku_list.append(sku_copy)

        if min_consign_price is None or min_market_price is None:
            raise ValueError('sku price parse failed: no valid sku prices')

        if modified:
            detail_data['sku_info'] = json.dumps(corrected_sku_list, ensure_ascii=False)

        return {
            'min_consign_price': float(min_consign_price),
            'min_market_price': float(min_market_price),
        }

    def _safe_decimal_non_negative(self, value, default='0.00') -> Decimal:
        try:
            amount = Decimal(str(value if value is not None and value != '' else default))
        except Exception:
            amount = Decimal(default)
        if amount < 0:
            amount = Decimal('0.00')
        return amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    def _calculate_spu_cost_price(self, sell_price, freight) -> float:
        sell_amount = self._safe_decimal_non_negative(sell_price)
        return float((sell_amount * Decimal('1.7')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))

    def _calculate_sku_market_price(self, consign_price, freight) -> float:
        consign_amount = self._safe_decimal_non_negative(consign_price)
        return float((consign_amount * Decimal('1.7')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))

    def _parse_sku_list(self, sku_info_value):
        if isinstance(sku_info_value, list):
            return sku_info_value
        if isinstance(sku_info_value, str):
            try:
                parsed = json.loads(sku_info_value)
                return parsed if isinstance(parsed, list) else []
            except Exception:
                return []
        return []

    def _build_existing_sku_market_map(self, existing_product: Optional[Dict]) -> Dict[str, Optional[float]]:
        existing_sku_info = self._parse_sku_list((existing_product or {}).get('sku_info'))
        sku_market_map = {}
        for sku in existing_sku_info:
            if not isinstance(sku, dict):
                continue
            sku_id = str(sku.get('skuId') or sku.get('sku_id') or sku.get('specId') or sku.get('spec_id') or '').strip()
            if not sku_id:
                continue
            market_price = sku.get('marketPrice', sku.get('market_price'))
            sku_market_map[sku_id] = float(self._safe_decimal_non_negative(market_price)) if market_price not in (None, '') else None
        return sku_market_map

    def _prepare_price_revision(self, product: Dict, existing_product: Optional[Dict] = None):
        if not isinstance(product, dict):
            raise ValueError('商品数据格式错误')

        detail_data = dict(product.get('detail_data') or {})
        if product.get('sku_price_checked'):
            revised_cost_price = float(self._safe_decimal_non_negative(product.get('cost_price', 0)))
            revised_product = dict(product)
            revised_product['cost_price'] = revised_cost_price
            revised_product['detail_data'] = detail_data
            return revised_product, [
                {
                    'revision_job_id': '',
                    'offer_id': str(product.get('offer_id') or '').strip(),
                    'sku_id': None,
                    'entity_type': 'SPU',
                    'field_name': 'cost_price',
                    'old_value': existing_product.get('cost_price') if isinstance(existing_product, dict) else None,
                    'new_value': revised_cost_price,
                    'source_type': str(product.get('source', SOURCE_TYPE)),
                }
            ]

        revised_cost_price = self._calculate_spu_cost_price(product.get('sell_price', 0), 0)
        revised_product = dict(product)
        revised_product['cost_price'] = revised_cost_price

        sku_list = self._parse_sku_list(detail_data.get('sku_info'))
        existing_sku_market_map = self._build_existing_sku_market_map(existing_product)
        price_logs = [
            {
                'revision_job_id': '',
                'offer_id': str(product.get('offer_id') or '').strip(),
                'sku_id': None,
                'entity_type': 'SPU',
                'field_name': 'cost_price',
                'old_value': existing_product.get('cost_price') if isinstance(existing_product, dict) else None,
                'new_value': revised_cost_price,
                'source_type': str(product.get('source', SOURCE_TYPE)),
            }
        ]

        revised_sku_list = []
        for sku in sku_list:
            if not isinstance(sku, dict):
                revised_sku_list.append(sku)
                continue
            sku_copy = dict(sku)
            sku_id = str(sku_copy.get('skuId') or sku_copy.get('sku_id') or sku_copy.get('specId') or sku_copy.get('spec_id') or '').strip()
            consign_price = sku_copy.get('consignPrice', sku_copy.get('consign_price', 0))
            revised_market_price = self._calculate_sku_market_price(consign_price, 0)
            sku_copy['marketPrice'] = revised_market_price
            revised_sku_list.append(sku_copy)
            price_logs.append({
                'revision_job_id': '',
                'offer_id': str(product.get('offer_id') or '').strip(),
                'sku_id': sku_id or None,
                'entity_type': 'SKU',
                'field_name': 'marketPrice',
                'old_value': existing_sku_market_map.get(sku_id),
                'new_value': revised_market_price,
                'source_type': str(product.get('source', SOURCE_TYPE)),
            })

        detail_data['sku_info'] = json.dumps(revised_sku_list, ensure_ascii=False) if revised_sku_list else ''
        revised_product['detail_data'] = detail_data
        return revised_product, price_logs

    def _load_existing_products(self, cursor, offer_ids: List[str]) -> Dict[str, Dict[str, Any]]:
        normalized_offer_ids = [str(offer_id).strip() for offer_id in offer_ids if str(offer_id).strip()]
        if not normalized_offer_ids:
            return {}
        placeholders = ','.join([PLACEHOLDER for _ in normalized_offer_ids])
        cursor.execute(
            f'SELECT offer_id, cost_price, sku_info FROM import_product WHERE offer_id IN ({placeholders})',
            normalized_offer_ids
        )
        return {str(row['offer_id']).strip(): dict(row) for row in cursor.fetchall()}

    def _insert_price_revision_logs(self, cursor, price_logs: List[Dict[str, Any]], revision_job_id: str):
        if not price_logs:
            return 0

        values = []
        for item in price_logs:
            values.append((
                revision_job_id,
                item.get('offer_id'),
                item.get('sku_id'),
                item.get('entity_type'),
                item.get('field_name'),
                item.get('old_value'),
                item.get('new_value'),
                item.get('source_type', SOURCE_TYPE),
                datetime.now()
            ))

        cursor.executemany(
            f'''
                INSERT INTO price_revision_log
                (revision_job_id, offer_id, sku_id, entity_type, field_name, old_value, new_value, source_type, operation_time)
                VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})
            ''',
            values
        )
        return len(values)
    
    def _batch_import_to_database(self, products: List[Dict]) -> Dict[str, Any]:
        """批量导入商品到数据库
        
        Args:
            products: 商品数据列表
            
        Returns:
            导入结果统计
        """
        if not products:
            return {'success_count': 0, 'failed_details': []}

        conn = get_db()
        cursor = conn.cursor()
        revision_job_id = uuid4().hex

        try:
            logger.info("[Price Revision] Start import revision job=%s product_count=%s", revision_job_id, len(products))
            existing_product_map = self._load_existing_products(cursor, [product.get('offer_id') for product in products])
            success_count = 0
            price_logs = []

            for i in range(0, len(products), self.BATCH_SIZE):
                batch = products[i:i + self.BATCH_SIZE]
                batch_result = self._execute_batch_insert(
                    conn,
                    cursor,
                    batch,
                    existing_product_map=existing_product_map,
                    revision_job_id=revision_job_id
                )
                success_count += batch_result['success_count']
                price_logs.extend(batch_result['price_logs'])

            self._insert_price_revision_logs(cursor, price_logs, revision_job_id)
            conn.commit()
            logger.info("[Price Revision] Commit import revision job=%s success_count=%s log_count=%s", revision_job_id, success_count, len(price_logs))
            return {
                'success_count': success_count,
                'failed_details': [],
                'revision_job_id': revision_job_id,
                'price_revision_count': len(price_logs)
            }
        except Exception as e:
            conn.rollback()
            logger.exception("[Price Revision] Rollback import revision job=%s", revision_job_id)
            return {
                'success_count': 0,
                'failed_details': [{'offer_id': '', 'reason': str(e)}],
                'revision_job_id': revision_job_id,
                'error': str(e)
            }
        finally:
            conn.close()
    
    def _execute_batch_insert(self, conn, cursor, products: List[Dict], existing_product_map: Optional[Dict[str, Dict[str, Any]]] = None, revision_job_id: str = '') -> Dict[str, Any]:
        """执行批量插入（优化版：先准备数据，再批量执行SQL）
        
        Args:
            products: 一批商品数据
            
        Returns:
            批次插入结果
        """
        success_count = 0
        price_logs = []
        existing_product_map = existing_product_map or {}
        
        INSERT_SQL = f'''
            INSERT INTO import_product
            (offer_id, title, price, cost_price, sell_price, freight, image_url,
             supplier_name, sales_count, deliver_days, stock,
             sync_status, description, attributes, images, sku_info, sku_count,
             supplier_id, category_id, category_name, erp_category_id, erp_category_name,
             offer_url, comment_count, month_order_count, month_distribution_count, tags, listed_time, shop_name, source_type,
             create_time, main_video, seven_days_refunds, product_type, quality_level, reference_price, seller_login_id,
             product_sale_info, product_extend_infos, sale_limit_address, service_capabilities, official_logistics_sku_info, product_shipping_info)
            VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER},
                    {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER},
                    'pending',
                    {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER},
                    {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER},
                    {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER},
                    {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER},
                    {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})
            ON DUPLICATE KEY UPDATE
                title = VALUES(title),
                price = VALUES(price),
                cost_price = VALUES(cost_price),
                sell_price = VALUES(sell_price),
                freight = VALUES(freight),
                image_url = VALUES(image_url),
                supplier_name = VALUES(supplier_name),
                sales_count = VALUES(sales_count),
                deliver_days = VALUES(deliver_days),
                stock = VALUES(stock),
                category_id = IFNULL(NULLIF(VALUES(category_id), ''), category_id),
                category_name = IFNULL(NULLIF(VALUES(category_name), ''), category_name),
                erp_category_id = IFNULL(NULLIF(VALUES(erp_category_id), ''), erp_category_id),
                erp_category_name = IFNULL(NULLIF(VALUES(erp_category_name), ''), erp_category_name),
                offer_url = IFNULL(NULLIF(VALUES(offer_url), ''), offer_url),
                comment_count = VALUES(comment_count),
                month_order_count = VALUES(month_order_count),
                month_distribution_count = VALUES(month_distribution_count),
                tags = IFNULL(NULLIF(VALUES(tags), ''), tags),
                listed_time = IFNULL(NULLIF(VALUES(listed_time), ''), listed_time),
                shop_name = IFNULL(NULLIF(VALUES(shop_name), ''), shop_name),
                source_type = VALUES(source_type),
                create_time = VALUES(create_time),
                main_video = VALUES(main_video),
                seven_days_refunds = VALUES(seven_days_refunds),
                product_type = VALUES(product_type),
                quality_level = VALUES(quality_level),
                reference_price = VALUES(reference_price),
                seller_login_id = VALUES(seller_login_id),
                product_sale_info = VALUES(product_sale_info),
                product_extend_infos = VALUES(product_extend_infos),
                sale_limit_address = VALUES(sale_limit_address),
                service_capabilities = VALUES(service_capabilities),
                official_logistics_sku_info = VALUES(official_logistics_sku_info),
                product_shipping_info = VALUES(product_shipping_info),
                sync_status = 'pending'
        '''
        
        try:
            all_values = []
            for product in products:
                offer_id = str(product.get('offer_id') or '').strip()
                revised_product, product_logs = self._prepare_price_revision(product, existing_product_map.get(offer_id))
                for item in product_logs:
                    item['revision_job_id'] = revision_job_id
                
                mapped_category = self._resolve_erp_category(
                    conn, 
                    cursor,
                    revised_product.get('detail_data', {}).get('category_name', ''),
                    revised_product.get('category_name_override', ''),
                    revised_product.get('title', '')
                )
                
                erp_category_id = str(mapped_category.get('id')) if mapped_category else ''
                erp_category_name = mapped_category.get('name', '') if mapped_category else ''
                
                detail_data = revised_product.get('detail_data', {})
                source_category_id = detail_data.get('category_id', '')
                source_category_name = detail_data.get('category_name', '')
                
                all_values.append((
                    revised_product['offer_id'],
                    revised_product['title'],
                    revised_product['price'],
                    revised_product['cost_price'],
                    revised_product['sell_price'],
                    revised_product.get('freight', 0),
                    revised_product['image_url'],
                    revised_product['supplier_name'],
                    revised_product['sales_count'],
                    revised_product['deliver_days'],
                    revised_product.get('stock', 0),
                    detail_data.get('description', ''),
                    detail_data.get('attributes', ''),
                    detail_data.get('images', ''),
                    detail_data.get('sku_info', ''),
                    detail_data.get('sku_count', 0),
                    detail_data.get('supplier_id', ''),
                    source_category_id,
                    source_category_name,
                    erp_category_id,
                    erp_category_name,
                    revised_product['offer_url'],
                    revised_product['comment_count'],
                    revised_product['month_order_count'],
                    revised_product['month_distribution_count'],
                    revised_product['tags'],
                    revised_product['listed_time'],
                    revised_product.get('shop_name', ''),
                    str(revised_product.get('source', SOURCE_TYPE)),
                    detail_data.get('create_time'),
                    detail_data.get('main_video'),
                    1 if detail_data.get('seven_days_refunds') else 0,
                    detail_data.get('product_type'),
                    detail_data.get('quality_level'),
                    detail_data.get('reference_price'),
                    detail_data.get('seller_login_id'),
                    detail_data.get('product_sale_info'),
                    detail_data.get('product_extend_infos'),
                    detail_data.get('sale_limit_address'),
                    detail_data.get('service_capabilities'),
                    detail_data.get('official_logistics_sku_info'),
                    detail_data.get('product_shipping_info')
                ))

                existing_product_map[offer_id] = {
                    'offer_id': offer_id,
                    'cost_price': revised_product['cost_price'],
                    'sku_info': detail_data.get('sku_info', '')
                }
                success_count += 1
                price_logs.extend(product_logs)
            
            if all_values:
                cursor.executemany(INSERT_SQL, all_values)
        except Exception:
            raise
        
        return {
            'success_count': success_count,
            'price_logs': price_logs
        }

    def rollback_price_revision(self, revision_job_id: str) -> Dict[str, Any]:
        revision_job_id = str(revision_job_id or '').strip()
        if not revision_job_id:
            return {'success': False, 'error': 'revision_job_id required'}

        conn = get_db()
        cursor = conn.cursor()
        rollback_job_id = uuid4().hex

        try:
            logger.info("[Price Revision] Start rollback revision_job_id=%s rollback_job_id=%s", revision_job_id, rollback_job_id)
            cursor.execute(
                f'''
                    SELECT id, offer_id, sku_id, entity_type, field_name, old_value, new_value, source_type, rolled_back_at
                    FROM price_revision_log
                    WHERE revision_job_id = {PLACEHOLDER}
                    ORDER BY id ASC
                ''',
                (revision_job_id,)
            )
            logs = [dict(row) for row in cursor.fetchall()]
            if not logs:
                return {'success': False, 'error': '未找到对应的价格修正日志'}
            if all(item.get('rolled_back_at') for item in logs):
                return {'success': True, 'message': '该批次已回滚，无需重复操作', 'revision_job_id': revision_job_id, 'rollback_job_id': rollback_job_id, 'restored_spu_count': 0, 'restored_sku_count': 0}

            offer_ids = sorted({str(item.get('offer_id') or '').strip() for item in logs if str(item.get('offer_id') or '').strip()})
            placeholders = ','.join([PLACEHOLDER for _ in offer_ids])
            cursor.execute(
                f'SELECT offer_id, cost_price, sku_info FROM import_product WHERE offer_id IN ({placeholders})',
                offer_ids
            )
            product_map = {str(row['offer_id']).strip(): dict(row) for row in cursor.fetchall()}

            restored_spu_count = 0
            restored_sku_count = 0

            for offer_id in offer_ids:
                product = product_map.get(offer_id)
                if not product:
                    continue
                sku_list = self._parse_sku_list(product.get('sku_info'))
                sku_index_map = {}
                for sku in sku_list:
                    if not isinstance(sku, dict):
                        continue
                    sku_id = str(sku.get('skuId') or sku.get('sku_id') or sku.get('specId') or sku.get('spec_id') or '').strip()
                    if sku_id:
                        sku_index_map[sku_id] = sku

                related_logs = [item for item in logs if str(item.get('offer_id') or '').strip() == offer_id]
                spu_old_value = None
                for item in related_logs:
                    if item.get('entity_type') == 'SPU' and item.get('field_name') == 'cost_price':
                        spu_old_value = item.get('old_value')
                    elif item.get('entity_type') == 'SKU' and item.get('field_name') == 'marketPrice':
                        sku_id = str(item.get('sku_id') or '').strip()
                        if sku_id and sku_id in sku_index_map:
                            sku_index_map[sku_id]['marketPrice'] = float(self._safe_decimal_non_negative(item.get('old_value', 0)))
                            restored_sku_count += 1

                cursor.execute(
                    '''
                        UPDATE import_product
                        SET cost_price = %s,
                            sku_info = %s
                        WHERE offer_id = %s
                    ''',
                    (
                        float(self._safe_decimal_non_negative(spu_old_value, default='0.00')) if spu_old_value is not None else product.get('cost_price'),
                        json.dumps(sku_list, ensure_ascii=False) if sku_list else product.get('sku_info', ''),
                        offer_id
                    )
                )
                restored_spu_count += 1

            cursor.execute(
                '''
                    UPDATE price_revision_log
                    SET rolled_back_at = %s,
                        rollback_job_id = %s
                    WHERE revision_job_id = %s
                ''',
                (datetime.now(), rollback_job_id, revision_job_id)
            )
            conn.commit()
            logger.info("[Price Revision] Rollback committed revision_job_id=%s rollback_job_id=%s restored_spu=%s restored_sku=%s", revision_job_id, rollback_job_id, restored_spu_count, restored_sku_count)
            return {
                'success': True,
                'message': '价格修正已回滚',
                'revision_job_id': revision_job_id,
                'rollback_job_id': rollback_job_id,
                'restored_spu_count': restored_spu_count,
                'restored_sku_count': restored_sku_count
            }
        except Exception as e:
            conn.rollback()
            logger.exception("[Price Revision] Rollback failed revision_job_id=%s", revision_job_id)
            return {'success': False, 'error': str(e), 'revision_job_id': revision_job_id}
        finally:
            conn.close()
    
    def _insert_single_product(self, product: Dict) -> bool:
        """插入单个商品（备用方法）"""
        conn = get_db()
        cursor = conn.cursor()
        
        try:
            # 复用批量插入中的逻辑
            # ... 简化版
            conn.commit()
            return True
        except Exception:
            return False
        finally:
            conn.close()
    
    _erp_category_cache = {}
    _erp_category_title_cache = {}
    
    def _resolve_erp_category(
        self, 
        conn, 
        cursor, 
        detail_category_name: str,
        category_name_override: str,
        title: str
    ):
        """解析ERP类目（复用连接，带缓存）"""
        try:
            category_text = category_name_override or detail_category_name
            cache_key = f"{category_text}|{title}"
            if cache_key in self._erp_category_cache:
                return self._erp_category_cache[cache_key]
            
            candidates = []
            
            if category_text:
                category_text = str(category_text).replace('|', '/').replace('>', '/').replace('＞', '/')
                for part in category_text.split('/'):
                    part = str(part).strip()
                    if part:
                        candidates.append(part)
            
            for name in reversed(candidates):
                name_key = f"__name__{name}"
                if name_key in self._erp_category_title_cache:
                    result = self._erp_category_title_cache[name_key]
                    if result is not None:
                        self._erp_category_cache[cache_key] = result
                        return result
                    continue
                
                cursor.execute(
                    f'''
                        SELECT c.id, c.name
                        FROM erp_category c
                        LEFT JOIN erp_category child ON child.parentId = c.id
                        WHERE c.name = {PLACEHOLDER}
                          AND child.id IS NULL
                          AND c.parentId IS NOT NULL AND c.parentId != 0
                        ORDER BY c.sort ASC, c.id DESC
                        LIMIT 1
                    ''',
                    (name,)
                )
                row = cursor.fetchone()
                if row:
                    result = dict(row)
                    self._erp_category_title_cache[name_key] = result
                    self._erp_category_cache[cache_key] = result
                    return result
                else:
                    self._erp_category_title_cache[name_key] = None
            
            if title:
                title_key = f"__title__{title}"
                if title_key in self._erp_category_title_cache:
                    result = self._erp_category_title_cache[title_key]
                    self._erp_category_cache[cache_key] = result
                    return result
                
                cursor.execute(
                    '''
                        SELECT c.id, c.name
                        FROM erp_category c
                        LEFT JOIN erp_category child ON child.parentId = c.id
                        WHERE CHAR_LENGTH(c.name) >= 2
                          AND %s LIKE CONCAT("%%", c.name, "%%")
                          AND child.id IS NULL
                          AND c.parentId IS NOT NULL AND c.parentId != 0
                        ORDER BY CHAR_LENGTH(c.name) DESC, c.sort ASC, c.id DESC
                        LIMIT 1
                    ''',
                    (str(title),)
                )
                row = cursor.fetchone()
                if row:
                    result = dict(row)
                    self._erp_category_title_cache[title_key] = result
                    self._erp_category_cache[cache_key] = result
                    return result
                else:
                    self._erp_category_title_cache[title_key] = None
            
            self._erp_category_cache[cache_key] = None
            return None
        except Exception:
            return None
    
    def _extract_offer_id(self, value):
        """从值中提取offer_id"""
        if value is None:
            return ''
        value_str = str(value).strip()
        if not value_str:
            return ''
        digits = re.sub(r'\D', '', value_str)
        if len(digits) >= 6:
            return digits
        match = re.search(r'/offer/(\d+)\.html', value_str)
        if match:
            return match.group(1)
        return ''
    
    def _to_float_value(self, value, default=0.0):
        """转换为浮点数"""
        if value is None:
            return float(default)
        value_str = str(value).strip()
        if not value_str:
            return float(default)
        value_str = value_str.replace(',', '')
        try:
            return float(value_str)
        except Exception:
            return float(default)
    
    def _to_int_value(self, value, default=0):
        """转换为整数"""
        try:
            return int(float(self._to_float_value(value, default)))
        except Exception:
            return int(default)


# 全局实例
product_import_service = ProductImportService()
