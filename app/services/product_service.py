from app.models import get_db
from app.config import SOURCE_TYPE
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import json
import csv
import io
import os
import threading
import re
import logging
from uuid import uuid4
from app.utils import calculate_adjusted_price_with_freight_fixed_nine, get_auto_adjust_ratio_by_sell_price, calculate_auto_adjusted_price_with_sell_price_tiers

PLACEHOLDER = '%s'
logger = logging.getLogger(__name__)

class ProductService:
    def _resolve_erp_category(self, category_name='', title=''):
        conn = get_db()
        cursor = conn.cursor()
        try:
            candidates = []
            if category_name:
                category_text = str(category_name).replace('|', '/').replace('>', '/').replace('＞', '/')
                for part in category_text.split('/'):
                    part = str(part).strip()
                    if part:
                        candidates.append(part)
            for name in reversed(candidates):
                cursor.execute(
                    f'''
                        SELECT c.id, c.name
                        FROM erp_category c
                        LEFT JOIN erp_category child ON child.parentId = c.id
                        WHERE c.name = {PLACEHOLDER}
                          AND child.id IS NULL
                          AND c.parentId IS NOT NULL AND c.parentId != 0
                        ORDER BY c.sort ASC, c.id DESC
                        LIMIT 1
                    ''',
                    (name,)
                )
                row = cursor.fetchone()
                if row:
                    return dict(row)
            if title:
                cursor.execute(
                    '''
                        SELECT c.id, c.name
                        FROM erp_category c
                        LEFT JOIN erp_category child ON child.parentId = c.id
                        WHERE CHAR_LENGTH(c.name) >= 2
                          AND %s LIKE CONCAT("%%", c.name, "%%")
                          AND child.id IS NULL
                          AND c.parentId IS NOT NULL AND c.parentId != 0
                        ORDER BY CHAR_LENGTH(c.name) DESC, c.sort ASC, c.id DESC
                        LIMIT 1
                    ''',
                    (str(title),)
                )
                row = cursor.fetchone()
                if row:
                    return dict(row)
            return None
        finally:
            conn.close()

    def _safe_decimal_non_negative(self, value, default='0.00', field_name='', offer_id='', sku_id=''):
        try:
            if value is None or (isinstance(value, str) and value.strip() == ''):
                raise ValueError('empty')
            amount = Decimal(str(value).strip())
        except (InvalidOperation, ValueError, TypeError):
            amount = Decimal(str(default))
            if field_name:
                logger.warning(
                    "[Batch Adjust Price] Invalid %s detected, fallback to %s. offer_id=%s sku_id=%s raw_value=%r",
                    field_name,
                    amount,
                    offer_id,
                    sku_id or '',
                    value
                )

        if amount < 0:
            logger.warning(
                "[Batch Adjust Price] Negative %s detected, fallback to 0. offer_id=%s sku_id=%s raw_value=%r",
                field_name or 'amount',
                offer_id,
                sku_id or '',
                value
            )
            amount = Decimal('0.00')

        return amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    def _safe_decimal_optional(self, value):
        try:
            if value is None or (isinstance(value, str) and value.strip() == ''):
                return None
            return float(Decimal(str(value).strip()).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))
        except (InvalidOperation, ValueError, TypeError):
            return None

    def _calculate_market_price(self, base_price, freight, multiplier='1.7'):
        base_amount = self._safe_decimal_non_negative(base_price)
        multiplier_amount = self._safe_decimal_non_negative(multiplier, default='1.70')
        return float((base_amount * multiplier_amount).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))

    def _parse_sku_list_for_price_update(self, sku_info_value, offer_id):
        if sku_info_value in (None, ''):
            return []
        if isinstance(sku_info_value, list):
            return sku_info_value
        if isinstance(sku_info_value, str):
            try:
                parsed = json.loads(sku_info_value)
            except Exception as exc:
                raise ValueError(f'offer_id={offer_id} sku_info JSON parse failed: {exc}')
            if isinstance(parsed, list):
                return parsed
        raise ValueError(f'offer_id={offer_id} sku_info must be a JSON array')

    def _extract_sku_identifier(self, sku, fallback=''):
        if not isinstance(sku, dict):
            return str(fallback or '').strip()
        sku_id = sku.get('sku_id') or sku.get('skuId') or sku.get('spec_id') or sku.get('specId') or fallback
        return str(sku_id or '').strip()

    def add_to_library(self, offer_id, title, price, image_url, supplier_name,
                       cost_price=None, supplier_location=None, sales_count=0,
                       support_return=False, deliver_days=48, detail_data=None, sell_price_override=None, category_name_override='',
                       offer_url='', comment_count=0,
                       month_order_count=0, month_distribution_count=0,
                       tags='', listed_time='', shop_name='', source=SOURCE_TYPE):
        conn = get_db()
        cursor = conn.cursor()
        
        if cost_price is None:
            cost_price = price

        # 直接使用传入的售价，不做自动计算（保持Excel原始数据）
        sell_price = sell_price_override if sell_price_override is not None else 0
        
        try:
            detail_fetched = 1 if detail_data else 0
            source_category_id = detail_data.get('category_id', '') if detail_data else ''
            source_category_name = detail_data.get('category_name', '') if detail_data else ''
            if not source_category_name and category_name_override:
                source_category_name = str(category_name_override).strip()
            mapped_category = self._resolve_erp_category(source_category_name, title)
            erp_category_id = str(mapped_category.get('id')) if mapped_category else ''
            erp_category_name = mapped_category.get('name', '') if mapped_category else ''

            cursor.execute(f'''
                INSERT INTO import_product
                (offer_id, title, price, cost_price, sell_price, image_url,
                 supplier_name, sales_count, deliver_days, stock,
                 sync_status, description, attributes, images, sku_info, sku_count,
                 supplier_id, category_id, category_name, erp_category_id, erp_category_name,
                 offer_url, comment_count, month_order_count, month_distribution_count, tags, listed_time, shop_name, source_type)
                VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER},
                        {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER},
                        'pending',
                        {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER},
                        {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER},
                        {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})
                ON DUPLICATE KEY UPDATE
                    title = VALUES(title),
                    price = VALUES(price),
                    cost_price = VALUES(cost_price),
                    sell_price = VALUES(sell_price),
                    image_url = VALUES(image_url),
                    supplier_name = VALUES(supplier_name),
                    sales_count = VALUES(sales_count),
                    deliver_days = VALUES(deliver_days),
                    stock = VALUES(stock),
                    category_id = IFNULL(NULLIF(VALUES(category_id), ''), category_id),
                    category_name = IFNULL(NULLIF(VALUES(category_name), ''), category_name),
                    erp_category_id = IFNULL(NULLIF(VALUES(erp_category_id), ''), erp_category_id),
                    erp_category_name = IFNULL(NULLIF(VALUES(erp_category_name), ''), erp_category_name),
                    offer_url = IFNULL(NULLIF(VALUES(offer_url), ''), offer_url),
                    comment_count = VALUES(comment_count),
                    month_order_count = VALUES(month_order_count),
                    month_distribution_count = VALUES(month_distribution_count),
                    tags = IFNULL(NULLIF(VALUES(tags), ''), tags),
                    listed_time = IFNULL(NULLIF(VALUES(listed_time), ''), listed_time),
                    shop_name = IFNULL(NULLIF(VALUES(shop_name), ''), shop_name),
                    source_type = VALUES(source_type),
                    sync_status = 'pending'
            ''', (
                offer_id, title, price, cost_price, sell_price, image_url,
                supplier_name, sales_count, deliver_days,
                0,
                detail_data.get('description', '') if detail_data else '',
                detail_data.get('attributes', '') if detail_data else '',
                detail_data.get('images', '') if detail_data else '',
                detail_data.get('sku_info', '') if detail_data else '',
                detail_data.get('sku_count', 0) if detail_data else 0,
                detail_data.get('supplier_id', '') if detail_data else '',
                source_category_id,
                source_category_name,
                erp_category_id,
                erp_category_name,
                offer_url,
                comment_count,
                month_order_count,
                month_distribution_count,
                tags,
                listed_time,
                shop_name,
                str(source or SOURCE_TYPE)
            ))
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Add to library error: {e}")
            conn.close()
            return False

    def _extract_offer_id(self, value):
        if value is None:
            return ''
        value_str = str(value).strip()
        if not value_str:
            return ''
        digits = re.sub(r'\D', '', value_str)
        if len(digits) >= 6:
            return digits
        match = re.search(r'/offer/(\d+)\.html', value_str)
        if match:
            return match.group(1)
        return ''

    def _to_float_value(self, value, default=0.0):
        if value is None:
            return float(default)
        value_str = str(value).strip()
        if not value_str:
            return float(default)
        value_str = value_str.replace(',', '')
        try:
            return float(value_str)
        except Exception:
            return float(default)

    def _to_int_value(self, value, default=0):
        try:
            return int(float(self._to_float_value(value, default)))
        except Exception:
            return int(default)

    def import_products_from_excel(self, file_stream):
        from openpyxl import load_workbook
        from app.services.ali1688_service import ali1688_service

        workbook = load_workbook(filename=file_stream, data_only=True)
        sheet = workbook.active
        if sheet.max_row < 2:
            return {'success': False, 'error': 'Excel无有效数据行'}

        headers = [str(cell.value).strip() if cell.value is not None else '' for cell in sheet[1]]
        required = ['宝贝ID', '商品标题']
        missing = [item for item in required if item not in headers]
        if missing:
            return {'success': False, 'error': f"Excel缺少必要列: {','.join(missing)}"}

        success_count = 0
        yx_count = 0
        df_count = 0
        failed_details = []

        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_data = {headers[idx]: row[idx] for idx in range(len(headers))}
            offer_id = self._extract_offer_id(row_data.get('宝贝ID') or row_data.get('宝贝链接'))
            if not offer_id:
                failed_details.append({'row': row_index, 'offer_id': '', 'reason': '宝贝ID为空或格式错误'})
                continue
            title = str(row_data.get('商品标题') or '').strip() or f'商品{offer_id}'
            image_url = str(row_data.get('图片地址') or '').strip()
            supplier_name = str(row_data.get('店铺名') or row_data.get('供应商') or '').strip()
            supplier_location = str(row_data.get('发货地') or '').strip()
            price = self._to_float_value(row_data.get('价格'), 0.0)
            sell_price = self._to_float_value(row_data.get('代发价'), 0.0)
            sales_count = self._to_int_value(row_data.get('销量'), 0)
            deliver_days = self._to_int_value(row_data.get('发货时间'), 48)
            support_return = '退' in str(row_data.get('标签') or '')
            category_name = str(row_data.get('类目') or '').strip()
            offer_url = str(row_data.get('宝贝链接') or '').strip()
            comment_count = self._to_int_value(row_data.get('评论数'), 0)
            month_order_count = self._to_int_value(row_data.get('月成交笔数'), 0)
            month_distribution_count = self._to_int_value(row_data.get('月代销'), 0)
            tags = str(row_data.get('标签') or '').strip()
            listed_time = str(row_data.get('上架时间') or '').strip()
            shop_name = str(row_data.get('店铺') or '').strip()

            detail_result = ali1688_service.get_yx_product_detail_for_import(offer_id)
            source = SOURCE_TYPE
            if detail_result.get('is_df'):
                source = SOURCE_TYPE
                detail_result = ali1688_service.get_distribution_product_info(offer_id)
            if not detail_result.get('success'):
                failed_details.append({'row': row_index, 'offer_id': offer_id, 'reason': detail_result.get('error', '详情获取失败')})
                continue
            detail_data = detail_result.get('detail') if detail_result.get('detail') else {}
            if not isinstance(detail_data, dict):
                detail_data = {}
            if category_name and not detail_data.get('category_name'):
                detail_data['category_name'] = category_name

            ok = self.add_to_library(
                offer_id=offer_id,
                title=title,
                price=price,
                image_url=image_url,
                supplier_name=supplier_name,
                cost_price=price,
                supplier_location=supplier_location,
                sales_count=sales_count,
                support_return=support_return,
                deliver_days=deliver_days if deliver_days > 0 else 48,
                detail_data=detail_data,
                sell_price_override=sell_price,  # 直接使用Excel中的代发价，不做自动计算
                category_name_override=category_name,
                offer_url=offer_url,
                comment_count=comment_count,
                month_order_count=month_order_count,
                month_distribution_count=month_distribution_count,
                tags=tags,
                listed_time=listed_time,
                shop_name=shop_name,
                source=source
            )
            if ok:
                success_count += 1
                if source == SOURCE_TYPE:
                    df_count += 1
                else:
                    yx_count += 1
            else:
                failed_details.append({'row': row_index, 'offer_id': offer_id, 'reason': '入库失败'})

        total = max(sheet.max_row - 1, 0)
        fail_count = total - success_count
        return {
            'success': fail_count == 0,
            'partial_success': fail_count > 0 and success_count > 0,
            'total': total,
            'success_count': success_count,
            'yx_count': yx_count,
            'df_count': df_count,
            'fail_count': fail_count,
            'failed_details': failed_details
        }
    
    def get_product(self, offer_id, source=SOURCE_TYPE):
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM import_product WHERE offer_id = %s AND source_type = %s', (offer_id, source))
        row = cursor.fetchone()
        conn.close()

        if row:
            return dict(row)
        return None

    def get_product_by_id(self, product_id, source=SOURCE_TYPE, conn=None, for_update=False):
        managed_conn = conn or get_db()
        cursor = managed_conn.cursor()
        sql = 'SELECT * FROM import_product WHERE id = %s AND source_type = %s'
        if for_update:
            sql += ' FOR UPDATE'
        cursor.execute(sql, (product_id, source))
        row = cursor.fetchone()
        if conn is None:
            managed_conn.close()
        if row:
            return dict(row)
        return None

    def delete_product_by_id(self, product_id, source=SOURCE_TYPE, conn=None):
        managed_conn = conn or get_db()
        cursor = managed_conn.cursor()
        cursor.execute(
            '''
                DELETE FROM import_product
                WHERE id = %s AND source_type = %s
            ''',
            (product_id, source)
        )
        affected = cursor.rowcount
        if conn is None:
            managed_conn.commit()
            managed_conn.close()
        return affected

    def delete_product_by_id_transactional(self, product_id, source=SOURCE_TYPE, erp_delete_executor=None):
        conn = get_db()
        try:
            if hasattr(conn, 'begin'):
                conn.begin()

            product = self.get_product_by_id(product_id, source=source, conn=conn, for_update=True)
            if not product:
                conn.rollback()
                return {
                    'success': True,
                    'local_missing': True,
                    'message': '本地数据已不存在，无需重复删除'
                }

            sync_status = str(product.get('sync_status') or 'pending').strip().lower()
            erp_delete_result = None
            if sync_status == 'synced':
                if not callable(erp_delete_executor):
                    conn.rollback()
                    return {
                        'success': False,
                        'code': 'ERP_DELETE_EXECUTOR_REQUIRED',
                        'error': 'ERP 删除执行器未配置'
                    }
                erp_delete_result = erp_delete_executor(product) or {}
                if not erp_delete_result.get('success'):
                    conn.rollback()
                    logger.error(
                        "[Product Delete] ERP delete failed. product_id=%s offer_id=%s error=%s",
                        product_id,
                        product.get('offer_id'),
                        erp_delete_result.get('error')
                    )
                    return {
                        'success': False,
                        'code': erp_delete_result.get('code') or 'ERP_DELETE_SERVICE_ERROR',
                        'error': erp_delete_result.get('error') or 'ERP 服务异常，请稍后重试',
                        'product': product
                    }

            affected = self.delete_product_by_id(product_id, source=source, conn=conn)
            conn.commit()

            result = {
                'success': True,
                'affected': affected,
                'product': product
            }

            if sync_status == 'synced':
                if erp_delete_result and erp_delete_result.get('not_found'):
                    result['erp_not_found'] = True
                    result['message'] = 'ERP 侧已不存在，仅清除本地数据'
                else:
                    result['message'] = '删除成功'
                result['erp_delete_result'] = erp_delete_result or {}
            else:
                result['message'] = '删除成功'

            return result
        except Exception as exc:
            try:
                conn.rollback()
            except Exception:
                pass
            logger.exception(
                "[Product Delete] Transactional delete failed. product_id=%s source=%s",
                product_id,
                source
            )
            return {
                'success': False,
                'code': 'DELETE_FAILED',
                'error': str(exc) or '删除失败'
            }
        finally:
            conn.close()
    
    def get_product_sku_info(self, offer_id, source=SOURCE_TYPE):
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT sku_info FROM import_product WHERE offer_id = %s AND source_type = %s', (offer_id, source))
        row = cursor.fetchone()
        conn.close()
        if row:
            sku_info = row['sku_info']
            if isinstance(sku_info, str):
                try:
                    return json.loads(sku_info)
                except Exception:
                    return sku_info
            return sku_info
        return None

    def get_product_shipping_info(self, offer_id, source=SOURCE_TYPE):
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT product_shipping_info FROM import_product WHERE offer_id = %s AND source_type = %s', (offer_id, source))
        row = cursor.fetchone()
        conn.close()
        if row:
            shipping_info = row.get('product_shipping_info')
            if shipping_info:
                if isinstance(shipping_info, str):
                    try:
                        return json.loads(shipping_info)
                    except Exception:
                        return shipping_info
                return shipping_info
        return None

    def batch_check_shipping_info(self, offer_ids, source=SOURCE_TYPE):
        if not offer_ids:
            return []
        conn = get_db()
        cursor = conn.cursor()
        placeholders = ','.join(['%s' for _ in offer_ids])
        cursor.execute(
            f'SELECT offer_id, title, product_shipping_info FROM import_product WHERE offer_id IN ({placeholders}) AND source_type = %s',
            offer_ids + [source]
        )
        results = []
        for row in cursor.fetchall():
            shipping_info = row.get('product_shipping_info')
            has_shipping = bool(shipping_info and str(shipping_info).strip())
            results.append({
                'offer_id': row.get('offer_id'),
                'title': row.get('title', ''),
                'has_shipping_info': has_shipping
            })
        conn.close()
        return results

    def check_shipping_info_stats(self, source=SOURCE_TYPE):
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(
            'SELECT COUNT(*) as total, '
            'SUM(CASE WHEN product_shipping_info IS NOT NULL AND product_shipping_info != \'\' THEN 1 ELSE 0 END) as has_shipping '
            'FROM import_product WHERE sync_status = %s AND source_type = %s',
            ('synced', source)
        )
        row = cursor.fetchone()
        conn.close()
        total = row.get('total', 0) if row else 0
        has_shipping = row.get('has_shipping', 0) if row else 0
        return {
            'total': total,
            'has_shipping_count': has_shipping,
            'no_shipping_count': total - has_shipping
        }

    def get_synced_offer_ids(self, source=SOURCE_TYPE):
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(
            'SELECT offer_id FROM import_product WHERE sync_status = %s AND source_type = %s',
            ('synced', source)
        )
        offer_ids = [row.get('offer_id') for row in cursor.fetchall() if row.get('offer_id')]
        conn.close()
        return offer_ids

    def update_product_detail(self, offer_id, detail_data, source=SOURCE_TYPE):
        conn = get_db()
        cursor = conn.cursor()

        try:
            cursor.execute(
                f'SELECT title, category_id, category_name, erp_category_id, erp_category_name FROM import_product WHERE offer_id = %s AND source_type = %s',
                (offer_id, source)
            )
            existing = cursor.fetchone() or {}
            source_category_id = detail_data.get('category_id', '') or existing.get('category_id', '')
            source_category_name = detail_data.get('category_name', '') or existing.get('category_name', '')
            mapped_category = self._resolve_erp_category(
                source_category_name,
                existing.get('title', '')
            )
            erp_category_id = str(mapped_category.get('id')) if mapped_category else (existing.get('erp_category_id', '') or '')
            erp_category_name = mapped_category.get('name', '') if mapped_category else (existing.get('erp_category_name', '') or '')
            cursor.execute(f'''
                UPDATE import_product
                SET description = %s, attributes = %s, images = %s,
                    sku_info = %s, sku_count = %s,
                    supplier_id = %s, supplier_name = %s,
                    category_id = %s, category_name = %s,
                    erp_category_id = %s, erp_category_name = %s,
                    product_shipping_info = %s
                WHERE offer_id = %s AND source_type = %s
            ''', (
                detail_data.get('description', ''),
                detail_data.get('attributes', ''),
                detail_data.get('images', ''),
                detail_data.get('sku_info', ''),
                detail_data.get('sku_count', 0),
                detail_data.get('supplier_id', ''),
                detail_data.get('supplier_name', ''),
                source_category_id,
                source_category_name,
                erp_category_id,
                erp_category_name,
                detail_data.get('product_shipping_info', ''),
                offer_id,
                source
            ))
            conn.commit()
            success = cursor.rowcount > 0
            conn.close()
            return success
        except Exception as e:
            print(f"Update product detail error: {e}")
            conn.close()
            return False
    
    def list_products(self, status='all', page=1, page_size=20, filters=None, source=SOURCE_TYPE):
        conn = get_db()
        cursor = conn.cursor()

        offset = (page - 1) * page_size

        where_clauses = ['source_type = %s']
        params = [source]

        if filters:
            if filters.get('keyword'):
                where_clauses.append('(title LIKE %s OR supplier_name LIKE %s OR shop_name LIKE %s)')
                keyword_pattern = f"%{filters['keyword']}%"
                params.extend([keyword_pattern, keyword_pattern, keyword_pattern])

            if filters.get('include_keywords'):
                for kw in filters['include_keywords']:
                    where_clauses.append('title LIKE %s')
                    params.append(f'%{kw}%')

            if filters.get('exclude_keywords'):
                for kw in filters['exclude_keywords']:
                    where_clauses.append('title NOT LIKE %s')
                    params.append(f'%{kw}%')

            if filters.get('price_min') is not None:
                where_clauses.append('price >= %s')
                params.append(filters['price_min'])

            if filters.get('price_max') is not None:
                where_clauses.append('price <= %s')
                params.append(filters['price_max'])

            if filters.get('sales_min') is not None:
                where_clauses.append('sales_count >= %s')
                params.append(filters['sales_min'])

            if filters.get('sales_max') is not None:
                where_clauses.append('sales_count <= %s')
                params.append(filters['sales_max'])

            if filters.get('supplier'):
                where_clauses.append('supplier_name LIKE %s')
                params.append(f"%{filters['supplier']}%")

            if filters.get('location'):
                where_clauses.append('supplier_location LIKE %s')
                params.append(f"%{filters['location']}%")

            if filters.get('sync_status'):
                where_clauses.append('sync_status = %s')
                params.append(filters['sync_status'])

            price_adjusted = filters.get('price_adjusted')
            if price_adjusted == 'yes':
                where_clauses.append('adjusted_price IS NOT NULL AND adjusted_price > 0')
            elif price_adjusted == 'no':
                where_clauses.append('(adjusted_price IS NULL OR adjusted_price <= 0)')

        where_sql = ' AND '.join(where_clauses)

        list_fields = [
            'id', 'offer_id', 'title', 'price', 'cost_price', 'sell_price', 'adjusted_price',
            'freight', 'image_url', 'supplier_name', 'shop_name', 'sales_count', 'deliver_days',
            'stock', 'sync_status', 'sync_at', 'sync_error',
            'category_id', 'category_name', 'erp_category_id', 'erp_category_name',
            'offer_url', 'comment_count', 'month_order_count', 'month_distribution_count',
            'tags', 'listed_time', 'sku_count', 'source_type', 'create_time'
        ]
        field_list = ', '.join(list_fields)

        count_sql = f'SELECT COUNT(*) as total_count FROM import_product WHERE {where_sql}'
        cursor.execute(count_sql, params)
        total = cursor.fetchone()['total_count']

        data_sql = f'''
            SELECT {field_list}
            FROM import_product
            WHERE {where_sql}
            ORDER BY id DESC
            LIMIT %s OFFSET %s
        '''
        params.extend([page_size, offset])
        cursor.execute(data_sql, params)

        rows = cursor.fetchall()

        products = []
        for row in rows:
            product = dict(row)
            product['detail_url'] = f"https://detail.1688.com/offer/{product['offer_id']}.html"
            products.append(product)

        conn.close()

        return {
            'success': True,
            'products': products,
            'total': total,
            'page': page,
            'page_size': page_size
        }
    
    def update_status(self, offer_id, status, source=SOURCE_TYPE):
        conn = get_db()
        cursor = conn.cursor()

        listed_at = datetime.now().isoformat() if status == 'listed' else None

        cursor.execute('''
            UPDATE import_product
            SET status = %s, listed_at = %s
            WHERE offer_id = %s AND source_type = %s
        ''', (status, listed_at, offer_id, source))

        success = cursor.rowcount > 0
        conn.commit()
        conn.close()
        return success
    
    def update_price(self, offer_id, sell_price, source=SOURCE_TYPE):
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE import_product
            SET sell_price = %s
            WHERE offer_id = %s AND source_type = %s
        ''', (sell_price, offer_id, source))

        success = cursor.rowcount > 0
        conn.commit()
        conn.close()
        return success

    def _batch_update_adjusted_price_legacy(self, offer_ids, adjust_ratio, source=SOURCE_TYPE):
        if not offer_ids:
            return {'success': False, 'error': 'offer_ids required', 'affected': 0, 'total': 0}

        conn = get_db()
        cursor = conn.cursor()
        ratio = float(adjust_ratio)
        placeholders = ','.join(['%s' for _ in offer_ids])

        try:
            cursor.execute(f'''
                SELECT offer_id, sell_price, cost_price, adjusted_price, freight, sku_info
                FROM import_product
                WHERE offer_id IN ({placeholders}) AND source_type = %s
            ''', offer_ids + [source])
            products = cursor.fetchall()
            if not products:
                return {'success': False, 'error': 'products not found', 'affected': 0, 'total': 0}

            affected = 0
            details = []

            for product in products:
                offer_id = product['offer_id']
                sell_price = round(float(product.get('sell_price') or 0), 2)
                base_cost_price = float(product.get('cost_price') or 0)
                existing_adjusted_price = float(product.get('adjusted_price') or 0)
                adjusted_base_price = sell_price if sell_price > 0 else existing_adjusted_price
                adjusted_price = float(calculate_adjusted_price_with_freight_fixed_nine(adjusted_base_price, 0, ratio))
                market_price = round((base_cost_price if base_cost_price > 0 else adjusted_base_price) * 1.5, 2)
                sku_info_str = product.get('sku_info') or ''
                updated_sku_info = None

                if sku_info_str:
                    try:
                        sku_list = json.loads(sku_info_str) if isinstance(sku_info_str, str) else sku_info_str
                        if isinstance(sku_list, list):
                            for sku in sku_list:
                                base_price = float(sku.get('consign_price') or sku.get('consignPrice') or sku.get('price') or sku.get('adjusted_price') or 0)
                                sku_adjusted_price = float(calculate_adjusted_price_with_freight_fixed_nine(base_price, 0, ratio))
                                sku_market_price = round(base_price * 1.5, 2)
                                sku['adjusted_price'] = sku_adjusted_price
                                sku['marketPrice'] = sku_market_price
                            updated_sku_info = json.dumps(sku_list, ensure_ascii=False)
                    except Exception as e:
                        print(f"Update adjusted sku info error: {e}")

                if updated_sku_info is not None:
                    cursor.execute('''
                        UPDATE import_product
                        SET sell_price = %s, adjusted_price = %s, cost_price = %s, sku_info = %s
                        WHERE offer_id = %s AND source_type = %s
                    ''', (sell_price, adjusted_price, market_price, updated_sku_info, offer_id, source))
                else:
                    cursor.execute('''
                        UPDATE import_product
                        SET sell_price = %s, adjusted_price = %s, cost_price = %s
                        WHERE offer_id = %s AND source_type = %s
                    ''', (sell_price, adjusted_price, market_price, offer_id, source))

                if cursor.rowcount > 0:
                    affected += 1
                details.append({
                    'offer_id': offer_id,
                    'sell_price': sell_price,
                    'adjusted_price': adjusted_price,
                    'cost_price': market_price
                })

            conn.commit()
            return {
                'success': True,
                'affected': affected,
                'total': len(products),
                'details': details
            }
        except Exception as e:
            conn.rollback()
            return {
                'success': False,
                'error': str(e),
                'affected': 0,
                'total': len(offer_ids)
            }
        finally:
            conn.close()

    def batch_update_adjusted_price(self, offer_ids, adjust_ratio, source=SOURCE_TYPE):
        if not offer_ids:
            return {'success': False, 'error': 'offer_ids required', 'affected': 0, 'total': 0}

        conn = get_db()
        cursor = conn.cursor()
        normalized_offer_ids = [str(offer_id).strip() for offer_id in offer_ids if str(offer_id).strip()]
        if not normalized_offer_ids:
            conn.close()
            return {'success': False, 'error': 'offer_ids required', 'affected': 0, 'total': 0}

        ratio = float(adjust_ratio)
        placeholders = ','.join([PLACEHOLDER for _ in normalized_offer_ids])
        try:
            cursor.execute(
                f'''
                    SELECT offer_id, sell_price, cost_price, adjusted_price, freight, sku_info
                    FROM import_product
                    WHERE offer_id IN ({placeholders}) AND source_type = %s
                    FOR UPDATE
                ''',
                normalized_offer_ids + [source]
            )
            products = cursor.fetchall()
            if not products:
                return {'success': False, 'error': 'products not found', 'affected': 0, 'total': 0}

            product_map = {str(product.get('offer_id') or '').strip(): product for product in products}
            missing_offer_ids = [offer_id for offer_id in normalized_offer_ids if offer_id not in product_map]
            if missing_offer_ids:
                raise ValueError(f'missing products or source mismatch: {",".join(missing_offer_ids)}')

            affected = 0
            details = []

            for offer_id in normalized_offer_ids:
                product = product_map[offer_id]
                sell_amount = self._safe_decimal_non_negative(
                    product.get('sell_price'),
                    field_name='sell_price',
                    offer_id=offer_id
                )
                adjusted_price = float(calculate_adjusted_price_with_freight_fixed_nine(sell_amount, 0, ratio))
                market_price = self._calculate_market_price(sell_amount, 0)
                raw_sku_info = product.get('sku_info')
                sku_list = self._parse_sku_list_for_price_update(raw_sku_info, offer_id)
                updated_sku_info = raw_sku_info if raw_sku_info not in (None, '') else ''

                if sku_list:
                    updated_sku_list = []
                    for sku_index, sku in enumerate(sku_list):
                        if not isinstance(sku, dict):
                            updated_sku_list.append(sku)
                            continue

                        sku_copy = dict(sku)
                        sku_id = self._extract_sku_identifier(sku_copy, fallback=f'index:{sku_index}')
                        sku_adjust_base_price = float(
                            sku_copy.get('consign_price')
                            or sku_copy.get('consignPrice')
                            or sku_copy.get('price')
                            or sku_copy.get('adjusted_price')
                            or 0
                        )
                        consign_amount = self._safe_decimal_non_negative(
                            sku_copy.get('consign_price', sku_copy.get('consignPrice')),
                            field_name='consignPrice',
                            offer_id=offer_id,
                            sku_id=sku_id
                        )
                        sku_adjusted_price = float(
                            calculate_adjusted_price_with_freight_fixed_nine(sku_adjust_base_price, 0, ratio)
                        )
                        sku_market_price = self._calculate_market_price(consign_amount, 0)
                        sku_copy['adjusted_price'] = sku_adjusted_price
                        sku_copy['marketPrice'] = sku_market_price
                        updated_sku_list.append(sku_copy)

                    updated_sku_info = json.dumps(updated_sku_list, ensure_ascii=False)

                cursor.execute(
                    '''
                        UPDATE import_product
                        SET sell_price = %s, adjusted_price = %s, cost_price = %s, sku_info = %s
                        WHERE offer_id = %s AND source_type = %s
                    ''',
                    (float(sell_amount), adjusted_price, market_price, updated_sku_info, offer_id, source)
                )

                affected += 1
                details.append({
                    'offer_id': offer_id,
                    'sell_price': float(sell_amount),
                    'adjusted_price': adjusted_price,
                    'cost_price': market_price,
                    'sku_count': len(sku_list) if isinstance(sku_list, list) else 0
                })

            conn.commit()
            logger.info(
                "[Batch Adjust Price] Local market price updated without revision log. source=%s affected=%s",
                source,
                affected
            )
            return {
                'success': True,
                'affected': affected,
                'total': len(products),
                'details': details
            }
        except Exception as e:
            conn.rollback()
            logger.exception(
                "[Batch Adjust Price] Local market price update rolled back. source=%s offer_count=%s",
                source,
                len(normalized_offer_ids)
            )
            return {
                'success': False,
                'error': str(e),
                'affected': 0,
                'total': len(normalized_offer_ids)
            }
        finally:
            conn.close()

    def batch_auto_adjust_price_by_sell_price(self, offer_ids, source=SOURCE_TYPE):
        if not offer_ids:
            return {'success': False, 'error': 'offer_ids required', 'affected': 0, 'total': 0}

        conn = get_db()
        cursor = conn.cursor()
        normalized_offer_ids = [str(offer_id).strip() for offer_id in offer_ids if str(offer_id).strip()]
        placeholders = ','.join([PLACEHOLDER for _ in normalized_offer_ids])
        revision_job_id = uuid4().hex

        try:
            cursor.execute(
                f'''
                    SELECT offer_id, sell_price, cost_price, adjusted_price, freight, sku_info, stock
                    FROM import_product
                    WHERE offer_id IN ({placeholders}) AND source_type = %s
                    FOR UPDATE
                ''',
                normalized_offer_ids + [source]
            )
            products = cursor.fetchall()
            if not products:
                return {'success': False, 'error': '未找到商品数据', 'affected': 0, 'total': 0}

            price_logs = []
            details = []
            affected = 0
            success_count = 0
            fail_count = 0

            for product in products:
                offer_id = str(product.get('offer_id') or '').strip()
                sell_price = float(product.get('sell_price') or 0)
                stock = int(product.get('stock') or 0)
                current_adjusted_price = float(product.get('adjusted_price') or 0)
                ratio = get_auto_adjust_ratio_by_sell_price(sell_price)

                if stock < 0:
                    fail_count += 1
                    details.append({
                        'offer_id': offer_id,
                        'success': False,
                        'reason': '库存锁定失败，库存值异常',
                        'sell_price': sell_price
                    })
                    continue

                if ratio is None:
                    fail_count += 1
                    details.append({
                        'offer_id': offer_id,
                        'success': False,
                        'reason': 'sell_price 必须大于 0 才能执行自动调价',
                        'sell_price': sell_price
                    })
                    continue

                adjusted_price = float(calculate_auto_adjusted_price_with_sell_price_tiers(sell_price, 0, sell_price))
                sku_info_str = product.get('sku_info') or ''
                updated_sku_info = None
                sku_list = []

                if sku_info_str:
                    try:
                        sku_list = json.loads(sku_info_str) if isinstance(sku_info_str, str) else sku_info_str
                    except Exception:
                        sku_list = []

                if isinstance(sku_list, list):
                    for sku in sku_list:
                        if not isinstance(sku, dict):
                            continue
                        base_price = float(sku.get('consign_price') or sku.get('consignPrice') or sku.get('price') or sku.get('adjusted_price') or 0)
                        old_sku_adjusted_price = float(sku.get('adjusted_price') or 0)
                        new_sku_adjusted_price = float(calculate_auto_adjusted_price_with_sell_price_tiers(base_price, 0, sell_price))
                        sku['adjusted_price'] = new_sku_adjusted_price
                        sku_id = str(sku.get('sku_id') or sku.get('skuId') or sku.get('spec_id') or sku.get('specId') or '').strip() or None
                        price_logs.append((
                            revision_job_id,
                            None,
                            offer_id,
                            sku_id,
                            'SKU',
                            'adjusted_price',
                            old_sku_adjusted_price,
                            new_sku_adjusted_price,
                            source,
                            datetime.now(),
                            None
                        ))
                    updated_sku_info = json.dumps(sku_list, ensure_ascii=False)

                cursor.execute(
                    '''
                        UPDATE import_product
                        SET adjusted_price = %s, sku_info = %s
                        WHERE offer_id = %s AND source_type = %s
                    ''',
                    (adjusted_price, updated_sku_info if updated_sku_info is not None else sku_info_str, offer_id, source)
                )

                price_logs.append((
                    revision_job_id,
                    None,
                    offer_id,
                    None,
                    'SPU',
                    'adjusted_price',
                    current_adjusted_price,
                    adjusted_price,
                    source,
                    datetime.now(),
                    None
                ))
                affected += 1
                success_count += 1
                details.append({
                    'offer_id': offer_id,
                    'success': True,
                    'sell_price': sell_price,
                    'ratio': float(ratio),
                    'adjusted_price': adjusted_price,
                    'sku_count': len(sku_list) if isinstance(sku_list, list) else 0
                })

            if price_logs:
                cursor.executemany(
                    f'''
                        INSERT INTO price_revision_log
                        (revision_job_id, rollback_job_id, offer_id, sku_id, entity_type, field_name, old_value, new_value, source_type, operation_time, rolled_back_at)
                        VALUES ({PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER}, {PLACEHOLDER})
                    ''',
                    price_logs
                )

            conn.commit()
            return {
                'success': True,
                'affected': affected,
                'total': len(products),
                'success_count': success_count,
                'fail_count': fail_count,
                'details': details,
                'revision_job_id': revision_job_id
            }
        except Exception as e:
            conn.rollback()
            return {
                'success': False,
                'error': str(e),
                'affected': 0,
                'total': len(normalized_offer_ids),
                'revision_job_id': revision_job_id
            }
        finally:
            conn.close()

    def rollback_adjusted_price_revision(self, revision_job_id, source=SOURCE_TYPE):
        revision_job_id = str(revision_job_id or '').strip()
        if not revision_job_id:
            return {'success': False, 'error': 'revision_job_id required'}

        conn = get_db()
        cursor = conn.cursor()
        rollback_job_id = uuid4().hex

        try:
            cursor.execute(
                '''
                    SELECT id, offer_id, sku_id, entity_type, field_name, old_value, rolled_back_at
                    FROM price_revision_log
                    WHERE revision_job_id = %s AND field_name = 'adjusted_price'
                    ORDER BY id ASC
                ''',
                (revision_job_id,)
            )
            logs = cursor.fetchall()
            if not logs:
                return {'success': False, 'error': '未找到对应的调价日志'}
            if all(item.get('rolled_back_at') for item in logs):
                return {'success': True, 'message': '该批次已回滚', 'revision_job_id': revision_job_id, 'rollback_job_id': rollback_job_id}

            offer_ids = sorted({str(item.get('offer_id') or '').strip() for item in logs if str(item.get('offer_id') or '').strip()})
            placeholders = ','.join([PLACEHOLDER for _ in offer_ids])
            cursor.execute(
                f'''
                    SELECT offer_id, adjusted_price, sku_info
                    FROM import_product
                    WHERE offer_id IN ({placeholders}) AND source_type = %s
                    FOR UPDATE
                ''',
                offer_ids + [source]
            )
            product_map = {str(row['offer_id']).strip(): dict(row) for row in cursor.fetchall()}

            restored_count = 0
            for offer_id in offer_ids:
                product = product_map.get(offer_id)
                if not product:
                    continue
                sku_list = []
                sku_info_str = product.get('sku_info') or ''
                if sku_info_str:
                    try:
                        sku_list = json.loads(sku_info_str) if isinstance(sku_info_str, str) else sku_info_str
                    except Exception:
                        sku_list = []
                related_logs = [item for item in logs if str(item.get('offer_id') or '').strip() == offer_id]
                spu_old_value = product.get('adjusted_price')
                if isinstance(sku_list, list):
                    sku_map = {}
                    for sku in sku_list:
                        if not isinstance(sku, dict):
                            continue
                        sku_id = str(sku.get('sku_id') or sku.get('skuId') or sku.get('spec_id') or sku.get('specId') or '').strip()
                        if sku_id:
                            sku_map[sku_id] = sku
                    for item in related_logs:
                        if item.get('entity_type') == 'SKU':
                            sku_id = str(item.get('sku_id') or '').strip()
                            if sku_id and sku_id in sku_map:
                                sku_map[sku_id]['adjusted_price'] = float(item.get('old_value') or 0)
                        elif item.get('entity_type') == 'SPU':
                            spu_old_value = item.get('old_value')

                cursor.execute(
                    '''
                        UPDATE import_product
                        SET adjusted_price = %s, sku_info = %s
                        WHERE offer_id = %s AND source_type = %s
                    ''',
                    (
                        float(spu_old_value or 0),
                        json.dumps(sku_list, ensure_ascii=False) if isinstance(sku_list, list) else sku_info_str,
                        offer_id,
                        source
                    )
                )
                restored_count += 1

            cursor.execute(
                '''
                    UPDATE price_revision_log
                    SET rolled_back_at = %s, rollback_job_id = %s
                    WHERE revision_job_id = %s AND field_name = 'adjusted_price'
                ''',
                (datetime.now(), rollback_job_id, revision_job_id)
            )
            conn.commit()
            return {
                'success': True,
                'message': '批量调价已回滚',
                'revision_job_id': revision_job_id,
                'rollback_job_id': rollback_job_id,
                'restored_count': restored_count
            }
        except Exception as e:
            conn.rollback()
            return {'success': False, 'error': str(e), 'revision_job_id': revision_job_id}
        finally:
            conn.close()
    
    def update_price_and_insurance(self, offer_id, sell_price, insurance_fee, source=SOURCE_TYPE):
        conn = get_db()
        cursor = conn.cursor()
        
        # 获取商品详情
        cursor.execute('SELECT sku_info FROM import_product WHERE offer_id = %s AND source_type = %s', (offer_id, source))
        row = cursor.fetchone()
        sku_info_updated = False
        
        if row and row.get('sku_info'):
            try:
                import json
                sku_info = json.loads(row['sku_info'])
                if isinstance(sku_info, list):
                    # 更新所有SKU的价格
                    for sku in sku_info:
                        sku['consign_price'] = float(sell_price)
                        sku['price'] = float(sell_price)
                        sku['consignPrice'] = float(sell_price)
                    
                    cursor.execute('''
                        UPDATE import_product
                        SET sell_price = %s, insurance_fee = %s, sku_info = %s
                        WHERE offer_id = %s AND source_type = %s
                    ''', (sell_price, insurance_fee, json.dumps(sku_info, ensure_ascii=False), offer_id, source))
                    sku_info_updated = True
            except Exception as e:
                print(f"Update SKU info error: {e}")
        
        if not sku_info_updated:
            cursor.execute('''
                UPDATE import_product
                SET sell_price = %s, insurance_fee = %s
                WHERE offer_id = %s AND source_type = %s
            ''', (sell_price, insurance_fee, offer_id, source))

        success = cursor.rowcount > 0
        conn.commit()
        conn.close()
        return success
    
    def update_stock(self, offer_id, stock, source=SOURCE_TYPE):
        conn = get_db()
        cursor = conn.cursor()
        
        # 获取商品详情
        cursor.execute('SELECT sku_info FROM import_product WHERE offer_id = %s AND source_type = %s', (offer_id, source))
        row = cursor.fetchone()
        sku_info_updated = False
        
        if row and row.get('sku_info'):
            try:
                import json
                sku_info = json.loads(row['sku_info'])
                if isinstance(sku_info, list):
                    # 更新所有SKU的库存
                    for sku in sku_info:
                        sku['stock'] = int(stock)
                        sku['amountOnSale'] = int(stock)
                    
                    cursor.execute('''
                        UPDATE import_product
                        SET stock = %s, sku_info = %s
                        WHERE offer_id = %s AND source_type = %s
                    ''', (stock, json.dumps(sku_info, ensure_ascii=False), offer_id, source))
                    sku_info_updated = True
            except Exception as e:
                print(f"Update SKU stock error: {e}")
        
        if not sku_info_updated:
            cursor.execute('''
                UPDATE import_product
                SET stock = %s
                WHERE offer_id = %s AND source_type = %s
            ''', (stock, offer_id, source))

        success = cursor.rowcount > 0
        conn.commit()
        conn.close()
        return success
    
    def batch_update_status(self, offer_ids, status, source=SOURCE_TYPE):
        conn = get_db()
        cursor = conn.cursor()

        listed_at = datetime.now().isoformat() if status == 'listed' else None
        placeholders = ','.join(['%s' for _ in offer_ids])

        cursor.execute(f'''
            UPDATE import_product
            SET status = %s, listed_at = %s
            WHERE offer_id IN ({placeholders}) AND source_type = %s
        ''', [status, listed_at] + offer_ids + [source])

        affected = cursor.rowcount
        conn.commit()
        conn.close()
        return affected
    
    def batch_delete(self, offer_ids, source=SOURCE_TYPE):
        conn = get_db()
        cursor = conn.cursor()

        placeholders = ','.join(['%s' for _ in offer_ids])

        cursor.execute(f'''
            DELETE FROM import_product
            WHERE offer_id IN ({placeholders}) AND source_type = %s
        ''', offer_ids + [source])

        affected = cursor.rowcount
        conn.commit()
        conn.close()
        return affected
    
    def get_stats(self, source=SOURCE_TYPE):
        conn = get_db()
        cursor = conn.cursor()

        cursor.execute('''
            SELECT
                COUNT(*) as total,
                COUNT(DISTINCT supplier_name) as suppliers,
                SUM(CASE WHEN sync_status = 'pending' THEN 1 ELSE 0 END) as pending_sync,
                SUM(CASE WHEN sync_status = 'synced' THEN 1 ELSE 0 END) as synced,
                SUM(CASE WHEN sync_status = 'failed' THEN 1 ELSE 0 END) as sync_failed,
                SUM(CASE WHEN sync_status = 'synced' AND adjusted_price IS NOT NULL AND adjusted_price > 0 THEN 1 ELSE 0 END) as synced_adjusted,
                SUM(CASE WHEN sync_status = 'synced' AND (adjusted_price IS NULL OR adjusted_price <= 0) THEN 1 ELSE 0 END) as synced_not_adjusted
            FROM import_product
            WHERE source_type = %s
        ''', (source,))

        row = cursor.fetchone()
        conn.close()

        return {
            'total': row['total'] or 0,
            'suppliers': row['suppliers'] or 0,
            'synced': row['synced'] or 0,
            'pending_sync': row['pending_sync'] or 0,
            'sync_failed': row['sync_failed'] or 0,
            'synced_adjusted': row['synced_adjusted'] or 0,
            'synced_not_adjusted': row['synced_not_adjusted'] or 0,
        }
    
    def get_products_for_sync(self, offer_ids=None, include_detail=True, source=SOURCE_TYPE):
        conn = get_db()
        cursor = conn.cursor()

        if offer_ids:
            placeholders = ','.join(['%s' for _ in offer_ids])
            cursor.execute(f'SELECT * FROM import_product WHERE offer_id IN ({placeholders}) AND source_type = %s', offer_ids + [source])
        else:
            cursor.execute('SELECT * FROM import_product WHERE (sync_status = %s OR sync_status IS NULL) AND source_type = %s', ('pending', source))

        products = [dict(row) for row in cursor.fetchall()]
        conn.close()

        for p in products:
            p['detail_url'] = f"https://detail.1688.com/offer/{p['offer_id']}.html"

            if include_detail:
                if p.get('attributes'):
                    try:
                        p['attributes'] = json.loads(p['attributes'])
                    except:
                        pass
                if p.get('images'):
                    try:
                        p['images'] = json.loads(p['images'])
                    except:
                        pass
                if p.get('sku_info'):
                    try:
                        p['sku_info'] = json.loads(p['sku_info'])
                    except:
                        pass

        return products
    
    def update_sync_status(self, offer_id, sync_status, sync_error=None, source=SOURCE_TYPE):
        conn = get_db()
        cursor = conn.cursor()

        sync_at = datetime.now().isoformat() if sync_status == 'synced' else None

        cursor.execute('''
            UPDATE import_product
            SET sync_status = %s, sync_at = %s, sync_error = %s
            WHERE offer_id = %s AND source_type = %s
        ''', (sync_status, sync_at, sync_error, offer_id, source))

        success = cursor.rowcount > 0
        conn.commit()
        conn.close()
        return success
    
    def batch_update_sync_status(self, offer_ids, sync_status, sync_error=None, source=SOURCE_TYPE):
        conn = get_db()
        cursor = conn.cursor()

        sync_at = datetime.now().isoformat() if sync_status == 'synced' else None
        placeholders = ','.join(['%s' for _ in offer_ids])

        cursor.execute(f'''
            UPDATE import_product
            SET sync_status = %s, sync_at = %s, sync_error = %s
            WHERE offer_id IN ({placeholders}) AND source_type = %s
        ''', [sync_status, sync_at, sync_error] + offer_ids + [source])

        affected = cursor.rowcount
        conn.commit()
        conn.close()
        return affected

    def batch_update_sync_status_with_reasons(self, offer_reason_map, sync_status, source=SOURCE_TYPE):
        normalized = []
        for offer_id, reason in (offer_reason_map or {}).items():
            offer_id_str = str(offer_id or '').strip()
            if not offer_id_str:
                continue
            normalized.append((sync_status, None, str(reason or '').strip(), offer_id_str, source))

        if not normalized:
            return 0

        conn = get_db()
        cursor = conn.cursor()
        cursor.executemany(
            '''
                UPDATE import_product
                SET sync_status = %s, sync_at = %s, sync_error = %s
                WHERE offer_id = %s AND source_type = %s
            ''',
            normalized
        )
        affected = cursor.rowcount
        conn.commit()
        conn.close()
        return affected

    def batch_adjust_price(self, offer_ids, adjust_rate, adjust_fee, source=SOURCE_TYPE):
        """
        批量调整商品价格
        adjust_rate: 调价比例，如 0.2 表示加价 20%，-0.1 表示降价 10%
        """
        import json
        
        conn = get_db()
        cursor = conn.cursor()

        results = []
        placeholders = ','.join(['%s' for _ in offer_ids])

        cursor.execute(f'''
            SELECT offer_id, sell_price, cost_price, sku_info FROM import_product
            WHERE offer_id IN ({placeholders}) AND source_type = %s
        ''', offer_ids + [source])

        products = cursor.fetchall()

        for product in products:
            offer_id = product['offer_id']
            old_price = float(product['sell_price'] or 0)
            cost_price = float(product['cost_price'] or 0)

            new_price = round(old_price * (1 + adjust_rate), 2)

            if new_price <= 0:
                results.append({
                    'offer_id': offer_id,
                    'success': False,
                    'error': '调价后价格为非正数',
                    'old_price': old_price,
                    'new_price': None
                })
                continue

            cursor.execute(f'''
                UPDATE import_product
                SET sell_price = %s
                WHERE offer_id = %s AND source_type = %s
            ''', (new_price, offer_id, source))

            sku_info_str = product.get('sku_info', '')
            if sku_info_str:
                try:
                    sku_list = json.loads(sku_info_str)
                    if isinstance(sku_list, list):
                        for sku in sku_list:
                            old_sku_price = 0
                            if 'salePrice' in sku:
                                old_sku_price = float(sku.get('salePrice', {}).get('value', 0) or 0)
                                sku['salePrice'] = {
                                    'value': str(round(old_sku_price * (1 + adjust_rate), 2)),
                                    'text': str(round(old_sku_price * (1 + adjust_rate), 2))
                                }
                            elif 'price' in sku:
                                old_sku_price = float(sku.get('price', 0) or 0)
                                sku['price'] = round(old_sku_price * (1 + adjust_rate), 2)
                            
                            if 'priceInfo' in sku and isinstance(sku['priceInfo'], dict):
                                for price_key, price_val in sku['priceInfo'].items():
                                    if isinstance(price_val, dict) and 'value' in price_val:
                                        old_val = float(price_val.get('value', 0) or 0)
                                        price_val['value'] = str(round(old_val * (1 + adjust_rate), 2))
                                        if 'text' in price_val:
                                            price_val['text'] = str(round(old_val * (1 + adjust_rate), 2))
                            
                            if 'discountPrice' in sku and isinstance(sku['discountPrice'], dict):
                                for price_key, price_val in sku['discountPrice'].items():
                                    if isinstance(price_val, dict) and 'value' in price_val:
                                        old_val = float(price_val.get('value', 0) or 0)
                                        price_val['value'] = str(round(old_val * (1 + adjust_rate), 2))
                        
                        new_sku_info = json.dumps(sku_list, ensure_ascii=False)
                        cursor.execute(f'''
                            UPDATE import_product
                            SET sku_info = %s
                            WHERE offer_id = %s AND source_type = %s
                        ''', (new_sku_info, offer_id, source))
                except (json.JSONDecodeError, TypeError) as e:
                    print(f"[Batch Adjust Price] Failed to parse sku_info for {offer_id}: {e}")

            results.append({
                'offer_id': offer_id,
                'success': True,
                'old_price': old_price,
                'new_price': new_price
            })

        conn.commit()
        conn.close()

        success_count = sum(1 for r in results if r.get('success'))
        return {
            'success': success_count > 0,
            'affected': success_count,
            'total': len(offer_ids),
            'details': results
        }
    
    def export_products(self, offer_ids=None, format_type='json', source=SOURCE_TYPE):
        conn = get_db()
        cursor = conn.cursor()

        if offer_ids:
            placeholders = ','.join(['%s' for _ in offer_ids])
            cursor.execute(f'SELECT * FROM import_product WHERE offer_id IN ({placeholders}) AND source_type = %s ORDER BY id DESC', offer_ids + [source])
        else:
            cursor.execute('SELECT * FROM import_product WHERE source_type = %s ORDER BY id DESC', (source,))

        products = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        for p in products:
            p['detail_url'] = f"https://detail.1688.com/offer/{p['offer_id']}.html"
            for key, value in p.items():
                if isinstance(value, datetime):
                    p[key] = value.isoformat()

        export_fieldnames = [
            'offer_id', 'title', 'price', 'cost_price', 'sell_price', 'freight',
            'supplier_name', 'shop_name', 'stock',
            'sync_status', 'sync_at',
            'category_id', 'category_name', 'erp_category_id', 'erp_category_name',
            'offer_url', 'comment_count', 'month_order_count', 'month_distribution_count',
            'tags', 'listed_time'
        ]
        
        if format_type == 'json':
            return json.dumps(products, ensure_ascii=False, indent=2), 'application/json', 'products.json'
        
        elif format_type == 'csv':
            output = io.StringIO()
            if products:
                writer = csv.DictWriter(output, fieldnames=export_fieldnames, extrasaction='ignore')
                writer.writeheader()
                writer.writerows(products)
            return output.getvalue(), 'text/csv', 'products.csv'
        
        elif format_type == 'excel':
            output = io.StringIO()
            if products:
                writer = csv.DictWriter(output, fieldnames=export_fieldnames, extrasaction='ignore')
                writer.writeheader()
                writer.writerows(products)
            return output.getvalue(), 'text/csv', 'products.csv'
        
        return products, 'application/json', 'products.json'

product_service = ProductService()
